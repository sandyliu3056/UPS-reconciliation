"""把倉儲計費工具真的開起來,照著 ① → ⑤ 走一遍,每一頁截圖。
對話框全部換成假的,所以不會卡住。要有畫面(Xvfb 也行)與 tkinter:

    xvfb-run -a python3 test/ui_warehouse_billing.py [截圖資料夾]

截圖要 Pillow;沒裝就只跑檢查不截圖。"""
import sys, pathlib, tempfile, json, os
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))
import warehouse_billing as wb
import tkinter as tk

OUT = pathlib.Path(sys.argv[1] if len(sys.argv) > 1 else str(pathlib.Path(tempfile.mkdtemp()) / "shots"))
OUT.mkdir(parents=True, exist_ok=True)
wb.DATA = pathlib.Path(tempfile.mkdtemp()) / "warehouse_billing.json"

# ---- stub dialogs --------------------------------------------------------
msgs = []
class FakeBox:
    def showwarning(self, t, m): msgs.append(("warn", t, m)); print("   [warn]", m.replace("\n", " ")[:120])
    def showerror(self, t, m): msgs.append(("error", t, m)); print("   [error]", m)
    def showinfo(self, t, m): msgs.append(("info", t, m)); print("   [info]", m.replace("\n", " ")[:80])
    def askyesno(self, t, m): msgs.append(("ask", t, m)); print("   [ask→yes]", m[:80]); return True
wb.messagebox = FakeBox()

def check(cond, msg):
    if not cond:
        raise AssertionError("FAILED: " + msg)
    print("  ok", msg)

def shot(name):
    root.update_idletasks(); root.update()
    try:
        from PIL import ImageGrab
        x, y = root.winfo_rootx(), root.winfo_rooty()
        w, h = root.winfo_width(), root.winfo_height()
        ImageGrab.grab(bbox=(x, y, x + w, y + h)).save(OUT / f"{name}.png")
        print("   shot", name)
    except Exception as e:
        os.system(f"xwd -root -silent | convert xwd:- {OUT}/{name}.png 2>/dev/null")
        print("   shot (fallback)", name, e.__class__.__name__)

def tab(i):
    app.nb.select(app.pages[i]); root.update()

def rows(tv):
    return [tv.item(i, "values") for i in tv.get_children()]

# ---- boot: empty file ----------------------------------------------------
root = tk.Tk()
app = wb.App(root)
root.update()
print("== boot (empty)")
check(len(app.pages) == 5, "five pages")
check([app.nb.tab(p, "text").strip() for p in app.pages] == list(app.tr("pages")), "tab titles from T")
check(app.edit_lv == "L1" and app.lv_box.get() == "Default", "a Default level exists and is the editing level")
check(app.c_level_box.get() == "Default", "blank client form defaults to the editing level")
check(app.rate_w["c20"][0].get() == "380", "rates page shows the level's rates")
check(app.rate_who.cget("text") == "Editing: Default", "rates page names the level being edited")
tab(0); shot("01_levels_empty_en")

# ---- levels page ---------------------------------------------------------
print("== levels")
app.lv_name.set("Level A"); app.on_add_level(); root.update()
check(app.edit_lv == "L2" and app.lv_box.get() == "Level A", "new level becomes the editing level, top bar follows")
check([r[0] for r in rows(app.lv_tv)] == ["Default", "Level A"], "level table lists both")
check(rows(app.lv_tv)[1][3] == "Editing" and rows(app.lv_tv)[0][3] == "", "state column marks the editing level")
app.lv_name.set("Level A"); app.on_add_level()
check(msgs[-1][2] == app.tr("lv_dup"), "duplicate name warned")
app.lv_name.set("Level B"); app.on_copy_level(); root.update()
check(app.edit_lv == "L3" and app.b.level_rate("L3", "c20") == 380.0, "copy makes L3 with A's rates")
app.lv_name.set("Level B (big)"); app.on_rename_level(); root.update()
check(app.lv_box.get() == "Level B (big)" and rows(app.lv_tv)[2][0] == "Level B (big)", "rename shows everywhere")
# pick a row in the table -> editing level switches
app.lv_tv.selection_set("L2"); root.update()
check(app.edit_lv == "L2" and app.lv_box.get() == "Level A", "clicking a level row switches the editing level")
# top-bar combobox -> editing level switches
app.lv_box.current(0); app.on_pick_level(); root.update()
check(app.edit_lv == "L1" and app.lv_tv.selection() == ("L1",), "top-bar box switches the editing level and the table follows")
app.lv_tv.selection_set("L2"); root.update()
shot("02_levels_en")

# ---- rates page edits the editing level ----------------------------------
print("== rates")
tab(2)
app.rate_w["c20"][0].set("400")   # trace -> on_rate
root.update()
check(app.b.level_rate("L2", "c20") == 400.0 and app.b.level_rate("L1", "c20") == 380.0, "typing a rate writes to the editing level only")
check(rows(app.lv_tv)[1][2].startswith(str(app.b.rates_set("L2"))), "level table 'rates set' cell updates live")
app.rate_w["stor"][0].set("")
check(app.b.level_rate("L2", "stor") is None and app.rate_w["stor"][4].cget("text") == app.tr("byquote"), "blank rate = no set price")
app.rate_w["stor"][0].set("0.4")
shot("03_rates_en")

# ---- clients page --------------------------------------------------------
print("== clients")
tab(1)
app.c_code.set("gen"); app.c_name.set("Geniqua Client"); app.c_contact.set("Amy"); app.c_note.set("test")
check(app.c_level_box.get() == "Level A", "form level box defaults to the editing level (Level A)")
app.on_add(); root.update()
check(app.cur == "GEN" and app.b.resolve("GEN") == ("L2", "ok"), "client added on Level A")
check(rows(app.cl_tv)[0][2] == "Level A", "client table shows the level name")
check(rows(app.lv_tv)[1][1] == "1", "level table client count updates")
app.c_code.set("nol"); app.c_name.set("No Level Yet"); app.c_level_box.current(0); app.on_add(); root.update()
check(app.b.resolve("NOL") == ("", "none") and rows(app.cl_tv)[1][2] == "Not set", "client without a level shows Not set")
check(app.status.cget("text") == app.tr("no_level"), "status bar says the client has no level")
# update: move NOL to Default
app.select("NOL"); app.c_level_box.current(1); app.on_update(); root.update()
check(app.b.resolve("NOL") == ("L1", "ok") and rows(app.cl_tv)[1][2] == "Default", "update reassigns the level")
# a client pointing at a deleted level
app.b.clients["NOL"]["level"] = "L9"; app.fill_clients(); app.select("NOL"); root.update()
check(rows(app.cl_tv)[1][2] == "L9 (level deleted)", "lost level is spelled out, not shown as unset")
check(app.c_level_box.get() == "L9 (level deleted)" and app.level_from_box() == "L9", "form keeps the lost id until changed")
app.on_update(); check(app.b.clients["NOL"]["level"] == "L9", "saving the form unchanged keeps the lost id")
app.select("GEN"); root.update()
shot("04_clients_en")

# ---- delete level guard --------------------------------------------------
print("== delete guard")
app.select_level("L2"); app.on_del_level()
check(msgs[-1][0] == "warn" and "GEN" in msgs[-1][2], "deleting a level in use is refused and names the client")
app.select_level("L3"); app.on_del_level(); root.update()
check("L3" not in app.b.levels and app.edit_lv == "L1", "unused level deleted, editing falls back to the first level")

# ---- calc page -----------------------------------------------------------
print("== calc")
app.select("GEN"); tab(3)
check(app.calc_lv.cget("text") == "Rate level: Level A", "calc page names the client's level")
app.it_box.current(wb.KEYS.index("c20")); app.q_var.set("2"); app.on_add_line(); root.update()
check(app.total_val.cget("text") == "$800.00", "2 x 400 on Level A")
app.order_var.set("SO-1"); app.pick_qty_var.set("3"); app.pick_weight_var.set("4.5")
app.flow_box.current(1); app.on_flow_change(); app.on_add_pick(); root.update()
check(app.total_val.cget("text") == "$803.24", "outbound pick priced by the level's weight bracket")
app.storage_cbm_var.set("10"); app.on_add_storage(); root.update()
check(app.b.qty("GEN", "stor") == 10.0 and app.total_val.cget("text") == "$807.24", "storage fee uses the level's CBM-day rate")
# change the level's rate on ③ -> ④ follows
app.select_level("L2"); app.rate_w["c20"][0].set("410"); root.update()
check(app.total_val.cget("text") == "$827.24", "changing the level's rate re-prices the client")
shot("05_calc_en")
# client with a lost level
app.select("NOL"); root.update()
check(app.calc_lv.cget("text") == app.tr("lost_level").format(id="L9"), "calc page explains a lost level")
check(app.total_val.cget("text") == "—", "no total without a usable level")
app.q_var.set("1"); app.on_add_line()
check(msgs[-1][2] == app.tr("need_level"), "adding a line without a level is refused")
app.order_var.set("SO-2"); app.pick_qty_var.set("1"); app.pick_weight_var.set("1"); app.on_add_pick()
check(msgs[-1][2] == app.tr("need_level"), "adding a pick without a level is refused")
shot("06_calc_lost_level_en")

# ---- bill page -----------------------------------------------------------
print("== bill")
app.select("GEN"); tab(4); root.update()
check(app.bill_lv.cget("text") == "Rate level: Level A", "bill page names the level")
check(rows(app.bill_tv)[-1][5] == "$827.24", "bill total matches calc")
shot("07_bill_en")
app.select("NOL"); root.update()
check(rows(app.bill_tv) == [], "no invoice lines without a usable level")
app.on_export_bill()
check(msgs[-1][2] == app.tr("need_level"), "invoice export refused without a level")

# ---- chinese -------------------------------------------------------------
print("== 中文")
app.lang_box.set("中文"); app.on_lang(); root.update()
check([app.nb.tab(p, "text").strip() for p in app.pages] == ["① 層級", "② 客戶", "③ 費率", "④ 計算", "⑤ 帳單"], "中文分頁")
check(app.L["l_editing"].cget("text") == "正在編輯" and rows(app.lv_tv)[1][3] == "正在編輯", "中文 editing label")
check(rows(app.cl_tv)[1][2] == "L9（層級已刪除）", "中文 lost-level label")
check(app.rate_who.cget("text") == "正在編輯：Level A", "中文 rates header")
app.select("GEN"); tab(3); root.update()
check(app.calc_lv.cget("text") == "費率層級：Level A", "中文 calc note")
tab(0); shot("08_levels_zh"); tab(1); shot("09_clients_zh"); tab(2); shot("10_rates_zh"); tab(3); shot("11_calc_zh"); tab(4); shot("12_bill_zh")

# ---- export price sheet for the editing level + save/reload ---------------
print("== export / save / reload")
out = pathlib.Path(tempfile.mkdtemp())
wb.filedialog.asksaveasfilename = lambda **k: str(out / k.get("initialfile", "x.xlsx"))
wb.filedialog.askopenfilename = lambda **k: str(out / "Geniqua_Price_Sheet_Level_A.xlsx")
app.select_level("L2"); app.on_export_price_template()
f = out / "Geniqua_Price_Sheet_Level_A.xlsx"
check(f.exists(), "price sheet exported with the level in the file name")
from openpyxl import load_workbook
w = load_workbook(f); ws = w.active
check(ws["A1"].value.endswith("Level A"), "price sheet title names the level")
c20 = [r for r in ws.iter_rows(values_only=True) if r[1] == "20GP"][0]
check(c20[2] == 410, "exported rate is the level's rate")
# import it into Default
app.select_level("L1"); app.on_import(); root.update()
check(app.b.level_rate("L1", "c20") == 410.0, "import lands on the editing level (Default)")
check(app.status.cget("text").startswith("已從報價表更新「Default」"), "import message names the level")
app.select("GEN"); app.on_export_bill()
inv = [p for p in out.glob("GQL-*.xlsx")]
check(len(inv) == 1, "invoice exported")
ws = load_workbook(inv[0]).active
check(ws["G2"].value == "Rate Level" and ws["H2"].value == "Level A", "invoice header carries the rate level")
app.save()
d = json.loads(wb.DATA.read_text(encoding="utf-8"))
check(set(d) == {"levels", "clients"} and d["clients"]["GEN"]["level"] == "L2", "saved file has the new layout")
root.destroy()

# ---- reboot from the saved file + from an old-layout file -----------------
print("== reboot")
root = tk.Tk(); app = wb.App(root); root.update()
check(app.edit_lv == "L1" and [r[0] for r in rows(app.lv_tv)] == ["Default", "Level A"], "levels reload")
check(app.nb.index(app.nb.select()) == 3, "opens on ④ when clients exist")
check(app.total_val.cget("text") == "$827.24", "GEN re-prices identically after reload")
root.destroy()
old = {"clients": {"OLD": {"name": "Old Client", "contact": "", "note": "",
       "rates": dict(wb.Book.fresh_rates(), c20=250.0), "qty": {"c20": 1}, "amt": {}, "picks": []}}}
wb.DATA.write_text(json.dumps(old), encoding="utf-8")
root = tk.Tk(); app = wb.App(root); root.update()
check(app.b.level_name(app.b.resolve("OLD")[0]) == "OLD" and app.total_val.cget("text") == "$250.00", "old per-client file migrates: level named after the client, total unchanged")
tab(0); shot("13_migrated_en")
root.destroy()
print("\nALL UI TESTS PASSED;", len(msgs), "dialogs stubbed")
