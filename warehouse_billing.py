"""
Warehouse Billing / 倉儲計費

Levels · Clients · Rates · Calculate · Invoice
層級 · 客戶 · 費率 · 計算 · 帳單

Rates hang off a rate level, not off a client. A level is one complete
set of rates; each client points at a level, and any number of clients
can share one. Same model as the UPS reconciliation tool.
費率掛在層級上，不掛在客戶上：一層是一整份單價表，客戶指到哪一層就用
哪一層的單價，好幾個客戶可以共用同一層。跟 UPS 對帳工具同一套做法。

Lines marked Varies / mark-up / waived on the price sheet carry no fixed
amount; type the amount straight into the calculation page.
報價表上寫 Varies / 加成 / waived 的幾項沒有固定金額，在計算頁直接填金額。

Import price sheet reads that xlsx directly.
python warehouse_billing.py
"""

import json
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from tkinter import font as tkfont
import ctypes
import math
import os
import sys

DATA = Path(__file__).with_name("warehouse_billing.json")

UI_THEMES = {
    "Crayon": dict(BG="#fdf4e6", PANEL="#fffdf6", TAB_BG="#f7dcae",
                   ACCENT="#a4522f", TEXT="#40301f", LINE="#c39d75",
                   TITLE="#8b4a2b", SELECT="#ffe6bd",
                   STRIPE="#faf0df",
                   RUN="#ffc861", RUN_HOVER="#ffd98d"),
    "Milk Tea": dict(BG="#cbb189", PANEL="#dbc6a4", TAB_BG="#bfa477",
                     ACCENT="#a9743f", TEXT="#3b2f23", LINE="#a98c63",
                     TITLE="#7a4f24", SELECT="#e3c79a",
                     STRIPE="#d3bb96",
                     RUN="#a9743f", RUN_HOVER="#8c5e30"),
    "Brown Gold": dict(BG="#f4ecdc", PANEL="#fffdf7", TAB_BG="#ffb500",
                       ACCENT="#351c15", TEXT="#2b1a10", LINE="#caa356",
                       TITLE="#351c15", SELECT="#ffdd8f",
                       STRIPE="#f7f1e4",
                       RUN="#351c15", RUN_HOVER="#1f100b"),
    "Sage": dict(BG="#c9d5c2", PANEL="#e0e8da", TAB_BG="#adbea4",
                 ACCENT="#4f6b48", TEXT="#263323", LINE="#93a889",
                 TITLE="#3c5236", SELECT="#cfe0c6",
                 STRIPE="#d6e0cf",
                 RUN="#4f6b48", RUN_HOVER="#3c5236"),
}


UI_SCALE = 1.0


def enable_dpi_awareness():
    """必須在建立 Tk 視窗之前呼叫，之後才宣告是沒有用的。"""
    if not sys.platform.startswith("win"):
        return
    try:
        ctypes.windll.user32.SetProcessDpiAwarenessContext(-4)
        return
    except Exception:
        pass
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
        return
    except Exception:
        pass
    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass


def measure_ui_scale(root):
    """算出這台螢幕的縮放倍率，並讓 Tk 照這個倍率排點數字體。

    UPS_UI_SCALE 環境變數可以強制指定，用來在別台機器上重現問題。
    """
    global UI_SCALE
    forced = os.environ.get("UPS_UI_SCALE", "").strip()
    dpi = None
    if forced:
        try:
            UI_SCALE = max(1.0, float(forced))
            dpi = 96.0 * UI_SCALE
        except ValueError:
            forced = ""
    if dpi is None:
        try:
            dpi = float(root.winfo_fpixels("1i"))
        except Exception:
            dpi = 96.0
        if sys.platform.startswith("win"):
            try:
                dpi = float(ctypes.windll.user32.GetDpiForSystem())
            except Exception:
                pass
        UI_SCALE = min(3.0, max(1.0, dpi / 96.0))
    try:
        root.tk.call("tk", "scaling", (96.0 * UI_SCALE) / 72.0)
    except tk.TclError:
        pass
    return UI_SCALE


def px(value):
    """寫死的像素值一律經過這裡。字體是點數會自己跟著縮放，畫布上的
    座標不會 —— 不乘回去，高解析度螢幕上按鈕會小得像郵票。"""
    return int(round(value * UI_SCALE))


UI_FONT_BASE = "Microsoft JhengHei UI"
FONT_BOOST = 0

ROUND_FONT_CANDIDATES = (
    "jf open 粉圓 1.1", "jf open 粉圓", "GenSenRounded TW", "源泉圓體",
    "GenJyuuGothic", "Zen Maru Gothic", "Yuanti TC", "圓體-繁",
    "Noto Sans TC", "思源黑體", "Source Han Sans TC",
    "Quicksand", "Poppins", "Nunito", "Varela Round",
)

HAND_FONT_CANDIDATES = (
    "芫荽", "Iansui", "jf open 手寫體", "華康少女文字W5",
    "GenSenRounded TW", "源泉圓體", "Zen Maru Gothic",
    "辰宇落雁體 2.0", "ChenYuluoyan 2.0", "辰宇落雁體", "Chenyuluoyan",
    "Ink Free", "Patrick Hand", "Architects Daughter", "Kalam",
    "Chalkboard SE", "Segoe Print",
)
HAND_CJK = {"芫荽", "Iansui", "jf open 手寫體", "華康少女文字W5",
            "GenSenRounded TW", "源泉圓體", "Zen Maru Gothic",
            "辰宇落雁體 2.0", "ChenYuluoyan 2.0", "辰宇落雁體",
            "Chenyuluoyan"}

FONT_TWEAKS = (
    ("chenyuluoyan", 3, True),
    ("落雁", 3, True),
    ("ink free", 1, False),
    ("segoe print", 1, False),
    ("iansui", 1, False),
    ("芫荽", 1, False),
)


def _tweak_for(family):
    name = (family or "").lower()
    for hint, step, bold in FONT_TWEAKS:
        if hint in name:
            return step, bold
    return 0, False


UI_FONT = UI_FONT_BASE
UI_FONT_DATA = UI_FONT_BASE
UI_FONT_TITLE = UI_FONT_BASE
UI_FONT_ROUND = None
UI_FONT_HAND = None

_FONT_POOL = {}


def _pick_font(root, candidates):
    try:
        installed = {name.strip() for name in tkfont.families(root)}
        for name in candidates:
            if name in installed:
                return name
    except Exception as error:
        print("Font lookup failed:", error)
    return None


def resolve_fonts(root, chinese_ui=False):
    """裝了哪一個圓體、哪一個手寫體就用哪一個。挑不到就講出來。"""
    global UI_FONT_ROUND, UI_FONT_HAND
    UI_FONT_ROUND = _pick_font(root, ROUND_FONT_CANDIDATES)
    UI_FONT_HAND = _pick_font(root, HAND_FONT_CANDIDATES)
    print("Round font:", UI_FONT_ROUND or "none installed")
    print("Hand font:", UI_FONT_HAND or "none installed")
    if not (UI_FONT_ROUND or UI_FONT_HAND):
        print("  -> 兩種都沒裝，介面只會有圓角沒有手寫字。"
              "裝一隻手寫字型就會自動生效。")
    hand = UI_FONT_HAND
    if hand and chinese_ui and hand not in HAND_CJK:
        hand = None
    set_ui_font(hand or UI_FONT_ROUND or UI_FONT_BASE,
                UI_FONT_ROUND or UI_FONT_BASE,
                hand or UI_FONT_ROUND or UI_FONT_BASE)


def set_ui_font(name, data_name=None, title_name=None):
    """三個角色的字體。不給就一路跟著主字體。"""
    global UI_FONT, UI_FONT_DATA, UI_FONT_TITLE
    UI_FONT = name
    UI_FONT_DATA = data_name or name
    UI_FONT_TITLE = title_name or name
    _refresh_font_pool()


def _family_for(role):
    return {"data": UI_FONT_DATA, "title": UI_FONT_TITLE}.get(role, UI_FONT)


def _role_spec(role, size, weight):
    family = _family_for(role)
    step, force_bold = _tweak_for(family)
    if force_bold and role != "data":
        weight = "bold"
    return dict(family=family, size=size + FONT_BOOST + step,
                weight=weight or "normal")


def _role_font(role, size, weight):
    key = (role, size, weight)
    font = _FONT_POOL.get(key)
    if font is None:
        spec = _role_spec(role, size, weight)
        try:
            font = tkfont.Font(**spec)
        except Exception:
            return (spec["family"], spec["size"], spec["weight"])
        _FONT_POOL[key] = font
    return font


def _refresh_font_pool():
    for (role, size, weight), font in _FONT_POOL.items():
        try:
            font.configure(**_role_spec(role, size, weight))
        except tk.TclError:
            pass
    apply_ui_fonts()
    refit_font_clients()


def ui_font(size=10, weight=None):
    """欄位標籤、說明、狀態列。"""
    return _role_font("ui", size, weight)


def data_font(size=10, weight=None):
    """清單、輸入框、下拉 —— 單號、金額、日期、路徑住在這裡。"""
    return _role_font("data", size, weight)


def title_font(size=10, weight=None):
    """分頁、區塊標題、按鈕。手寫的那一層，只用在短字串。"""
    return _role_font("title", size, weight)


_FONT_CLIENTS = []


def register_font_client(widget):
    _FONT_CLIENTS.append(widget)


def refit_font_clients():
    for widget in list(_FONT_CLIENTS):
        try:
            widget.refit()
        except tk.TclError:
            _FONT_CLIENTS.remove(widget)
        except Exception:
            pass


def _jitter(seed, index, amount):
    """固定的偏移量，不是亂數。

    每次重畫都重新亂數的話，滑鼠移過去邊框就會自己抖一下 —— 那是壞掉，
    不是手繪。同一顆按鈕的同一個點永遠偏同一個方向。
    """
    value = math.sin(seed * 12.9898 + index * 78.233) * 43758.5453
    return (value - math.floor(value) - 0.5) * 2.0 * amount


def sketch_rect_path(x0, y0, x1, y1, radius=10, seed=0.0, wobble=1.5):
    """畫歪的圓角矩形，回傳可以直接餵給 create_line / create_polygon 的座標。

    四個角各取幾個點畫弧，四條邊各補幾個點，每個點加一點偏移，再用
    smooth=True 連起來 —— 直線變成微微起伏的線，就是手畫的樣子。
    """
    radius = max(2, min(radius, (x1 - x0) / 2, (y1 - y0) / 2))
    points = []
    corners = (
        (x1 - radius, y1 - radius, 0),
        (x0 + radius, y1 - radius, 90),
        (x0 + radius, y0 + radius, 180),
        (x1 - radius, y0 + radius, 270),
    )
    index = 0
    for cx, cy, start in corners:
        for step in range(4):
            angle = math.radians(start + step * 30)
            px = cx + radius * math.cos(angle)
            py = cy + radius * math.sin(angle)
            points.append(px + _jitter(seed, index, wobble))
            points.append(py + _jitter(seed + 7.7, index, wobble))
            index += 1
    return points


def _mix(colour_a, colour_b, ratio):
    """把兩個顏色混起來。影子與第二道淺線都靠它，不必為了一層陰影
    在每一套配色裡多養一個顏色。"""
    try:
        first = colour_a.lstrip("#")
        second = colour_b.lstrip("#")
        parts = []
        for index in (0, 2, 4):
            one = int(first[index:index + 2], 16)
            two = int(second[index:index + 2], 16)
            parts.append(int(round(one + (two - one) * ratio)))
        return "#%02x%02x%02x" % tuple(parts)
    except Exception:
        return colour_a


def _lum(colour):
    """0（黑）到 255（白）。判斷字要白的還是墨的、點綴色跟底色夠不夠開。"""
    try:
        text = colour.lstrip("#")
        r, g, b = (int(text[i:i + 2], 16) for i in (0, 2, 4))
    except Exception:
        return 0.0
    return 0.299 * r + 0.587 * g + 0.114 * b


def _is_dark(colour):
    """深色底配白字，淺色底配墨字。四套配色的按鈕深淺差很多，
    字色寫死一種就會有一套看不見。"""
    return _lum(colour) < 150


POP = {
    "mint": "#7fb89a",
    "sky": "#7196bb",
    "coral": "#dd7f66",
    "sun": "#f0bf4a",
    "berry": "#c47b96",
}


def shade(colour, amount):
    """照 Reprice Tool 的 _mix：正數往白走，負數往黑走。"""
    try:
        value = int(str(colour).lstrip("#"), 16)
    except Exception:
        return colour
    r, g, b = (value >> 16) & 255, (value >> 8) & 255, value & 255
    if amount >= 0:
        r += (255 - r) * amount
        g += (255 - g) * amount
        b += (255 - b) * amount
    else:
        r *= 1 + amount
        g *= 1 + amount
        b *= 1 + amount
    return "#%02x%02x%02x" % (int(r), int(g), int(b))

class SketchButton(ttk.Button):
    """Reprice Tool 手繪殼的按鈕：ttk + clam，2px 描邊、圓角靠留白撐開。

    每一顆有自己的 style 名字 —— ttk 的樣式是照名字生效的，共用一個名字
    就沒辦法讓 Export 是薄荷綠、Clear 是珊瑚紅。
    """

    _count = 0

    def __init__(self, parent, text="", command=None, font=None, width=None,
                 padx=14, pady=7, tone="solid", pop=None,
                 page="#f4ecdc", fill="#ffb500", outline="#351c15",
                 hover="#351c15", fg=None, **ignored):
        SketchButton._count += 1
        self._style_name = "S%d.TButton" % SketchButton._count
        self._font = font or title_font(10, "bold")
        self._tone, self._pop = tone, pop
        self._pad = [px(padx), px(pady)]
        self._page, self._fill, self._outline = page, fill, outline
        self._hover, self._fg = hover, fg
        kwargs = {}
        if width:
            kwargs["width"] = width
        super().__init__(parent, text=text, command=command,
                         style=self._style_name, **kwargs)
        register_font_client(self)
        self._restyle()

    def retheme(self, page=None, fill=None, outline=None, hover=None,
                fg=None):
        self._page = page or self._page
        self._fill = fill or self._fill
        self._outline = outline or self._outline
        self._hover = hover or self._hover
        if fg is not None:
            self._fg = fg
        self._restyle()

    def refit(self):
        self._restyle()

    def _restyle(self):
        style = ttk.Style()
        if self._tone == "soft":
            edge = self._pop or self._outline
            if self._pop and abs(_lum(edge) - _lum(self._page)) < 58:
                edge = _mix(edge, self._fg or "#3b2f23", 0.34)
            background = self._page
            foreground = (_mix(edge, self._fg or "#3b2f23", 0.42)
                          if self._pop else (self._fg or self._outline))
            active_bg = _mix(edge, "#ffffff", 0.72)
        else:
            edge = self._outline
            background = self._fill
            foreground = self._fg or ("white" if _is_dark(self._fill)
                                      else self._outline)
            active_bg = self._hover
        try:
            style.configure(self._style_name, background=background,
                            foreground=foreground, bordercolor=edge,
                            lightcolor=background, darkcolor=background,
                            focuscolor=background, borderwidth=px(2),
                            relief="solid", padding=self._pad,
                            font=self._font, anchor="center")
            style.map(self._style_name,
                      background=[("pressed", active_bg),
                                  ("active", active_bg)],
                      foreground=[("pressed", foreground),
                                  ("active", foreground)],
                      bordercolor=[("active", edge)],
                      relief=[("pressed", "solid"), ("active", "solid")])
        except tk.TclError:
            pass

    def configure(self, **kwargs):
        """舊的呼叫端還可能用 bg / fg 這幾個名字進來，照收。"""
        for key, target in (("bg", "_fill"), ("background", "_fill"),
                            ("fg", "_fg"), ("foreground", "_fg"),
                            ("activebackground", "_hover")):
            if key in kwargs:
                setattr(self, target, kwargs.pop(key))
        kwargs.pop("activeforeground", None)
        if kwargs:
            ttk.Button.configure(self, **kwargs)
        self._restyle()

    config = configure


class SketchEntry(ttk.Entry):
    """Reprice Tool 手繪殼的輸入框：奶油色欄位、2px 邊、留白撐高。"""

    _count = 0

    def __init__(self, parent, textvariable=None, width=30, font=None,
                 page="#f4ecdc", panel="#fcfaf6", ink="#3b2f23",
                 line="#a98c63", **ignored):
        SketchEntry._count += 1
        self._style_name = "S%d.TEntry" % SketchEntry._count
        self._font = font or data_font(10)
        self._panel, self._ink, self._line = panel, ink, line
        super().__init__(parent, textvariable=textvariable, width=width,
                         style=self._style_name, font=self._font)
        register_font_client(self)
        self._restyle()

    def retheme(self, page=None, panel=None, ink=None, line=None, **ignored):
        self._panel = panel or self._panel
        self._ink = ink or self._ink
        self._line = line or self._line
        self._restyle()

    def refit(self):
        try:
            ttk.Entry.configure(self, font=self._font)
        except tk.TclError:
            pass

    def _restyle(self):
        try:
            ttk.Style().configure(self._style_name,
                                  fieldbackground=self._panel,
                                  foreground=self._ink,
                                  bordercolor=self._line,
                                  lightcolor=self._line, darkcolor=self._line,
                                  insertcolor=self._ink,
                                  borderwidth=px(2), relief="solid",
                                  padding=px(6))
        except tk.TclError:
            pass


class SketchCombo(ttk.Combobox):
    """Reprice Tool 手繪殼的下拉選單。

    唯讀的 combobox 會把值畫成「選取中的文字」，Tk 預設把它畫在藍色塊上
    —— 每一個狀態都設成跟欄位同色，那塊藍才會消失。
    """

    _count = 0

    def __init__(self, parent, textvariable=None, values=(), width=None,
                 font=None, page="#f4ecdc", panel="#fcfaf6", ink="#3b2f23",
                 line="#a98c63", hover="#e3c79a", **ignored):
        SketchCombo._count += 1
        self._style_name = "S%d.TCombobox" % SketchCombo._count
        self._font = font or title_font(10)
        self._panel, self._ink, self._line = panel, ink, line
        values = list(values)
        if width is None:
            width = max([len(str(v)) for v in values] or [10]) + 2
        super().__init__(parent, textvariable=textvariable, values=values,
                         width=width, state="readonly",
                         style=self._style_name, font=self._font)
        register_font_client(self)
        self._restyle()

    def retheme(self, page=None, panel=None, ink=None, line=None,
                hover=None, **ignored):
        self._panel = panel or self._panel
        self._ink = ink or self._ink
        self._line = line or self._line
        self._restyle()

    def refit(self):
        try:
            ttk.Combobox.configure(self, font=self._font)
        except tk.TclError:
            pass

    def _restyle(self):
        style = ttk.Style()
        try:
            style.configure(self._style_name, fieldbackground=self._panel,
                            background=self._panel, foreground=self._ink,
                            arrowcolor=self._ink, bordercolor=self._line,
                            lightcolor=self._line, darkcolor=self._line,
                            borderwidth=px(2), relief="solid",
                            padding=px(5), font=self._font)
            style.map(self._style_name,
                      fieldbackground=[("readonly", self._panel),
                                       ("disabled", self._panel)],
                      background=[("readonly", self._panel),
                                  ("active", self._panel)],
                      foreground=[("readonly", self._ink),
                                  ("disabled", self._line)],
                      selectbackground=[("readonly", self._panel),
                                        ("!focus", self._panel)],
                      selectforeground=[("readonly", self._ink),
                                        ("!focus", self._ink)],
                      arrowcolor=[("readonly", self._ink)])
            self.tk.call("option", "add", "*TCombobox*Listbox.background",
                         self._panel)
            self.tk.call("option", "add", "*TCombobox*Listbox.foreground",
                         self._ink)
            self.tk.call("option", "add", "*TCombobox*Listbox.font",
                         self._font)
        except tk.TclError:
            pass


class SketchPanel(tk.Frame):
    """清單外面那一圈。Reprice Tool 的殼是 2px 實線邊，不是畫歪的框。"""

    def __init__(self, parent, page="#f4ecdc", panel="#fcfaf6",
                 line="#a98c63", pad=2, **ignored):
        super().__init__(parent, bg=line, bd=0, highlightthickness=0)
        self._page, self._panel, self._line = page, panel, line
        self._pad = px(pad)
        self.body = tk.Frame(self, bg=panel, bd=0, highlightthickness=0)
        self.body.pack(fill="both", expand=True, padx=self._pad,
                       pady=self._pad)

    def retheme(self, page=None, panel=None, line=None):
        self._page = page or self._page
        self._panel = panel or self._panel
        self._line = line or self._line
        try:
            tk.Frame.configure(self, bg=self._line)
            self.body.configure(bg=self._panel)
        except tk.TclError:
            pass


class SketchScrollbar(ttk.Scrollbar):
    """ttk 捲軸，配色跟著主題走。

    介面上跟舊的那支相容的是 set() 與 command，這兩個就是捲動用得到的
    全部。
    """

    instances = []

    def __init__(self, parent, orient="vertical", command=None,
                 page="#f4ecdc", trough="#f7f1e4", line="#a98c63",
                 thumb="#bfa477", **ignored):
        name = "V" if orient == "vertical" else "H"
        SketchScrollbar.instances.append(self)
        self._style_name = "%s%d.%s.TScrollbar" % (
            name, len(SketchScrollbar.instances),
            "Vertical" if orient == "vertical" else "Horizontal")
        self._trough, self._line, self._thumb = trough, line, thumb
        super().__init__(parent, orient=orient, command=command,
                         style=self._style_name)
        self._restyle()

    def retheme(self, page=None, trough=None, line=None, thumb=None):
        self._trough = trough or self._trough
        self._line = line or self._line
        self._thumb = thumb or self._thumb
        self._restyle()

    def _restyle(self):
        try:
            ttk.Style().configure(self._style_name, background=self._thumb,
                                  troughcolor=self._trough,
                                  bordercolor=self._line,
                                  lightcolor=self._thumb,
                                  darkcolor=self._thumb,
                                  arrowcolor=self._line,
                                  borderwidth=px(1), relief="flat",
                                  arrowsize=px(12), width=px(13))
            ttk.Style().map(self._style_name,
                            background=[("active", self._line)])
        except tk.TclError:
            pass


class PetPainter:
    """畫 Reprice Tool 那對貓狗。座標與顏色照抄，只是畫布由外面給。

    材質色（毛色、鼻子、舌頭）寫死，跟 Reprice Tool 同一個做法 —— 主題管
    的是牆面、金屬、字，不是動物的毛色。
    """

    INK = "#241d17"
    CAT_COAT, CAT_MARK = "#f2ece1", "#cfc7b8"
    DOG_COAT, DOG_MARK = "#f5efe6", "#b5793f"

    def __init__(self, canvas, scale=1.0, line="#a98c63"):
        self.cv = canvas
        self.k = scale
        self.line = line

    def _S(self, value):
        return value * self.k

    @property
    def _floor(self):
        return shade(self.line, 0.30)

    def draw(self, kind, x, y, frame, facing=-1, walking=False):
        """畫一隻。facing=-1 是原本的朝向（朝左），+1 就整隻鏡射。

        走路一定要臉朝行進方向 —— 圖是朝左畫的，往右走不鏡射就是倒退嚕。
        """
        before = set(self.cv.find_all())
        self.floor_y = y
        stride = math.sin(frame * 0.55) if walking else 0.0
        bob = abs(math.sin(frame * 0.55)) * 0.9 if walking else 0.0
        if kind == "dog":
            self._doodle_dog(x, frame, y - bob, stride)
        else:
            self._doodle_cat(x, frame, y - bob, stride)
        if facing > 0:
            origin = self._S(x)
            for item in set(self.cv.find_all()) - before:
                self.cv.scale(item, origin, 0, -1, 1)

    def _blob(self, x0, y0, x1, y1, fill, w=2.2, ink=None, r=None):
        """圓角塊。Canvas 沒有圓角矩形 —— 用切角的多邊形加 smooth，轉角
        會被雲形曲線帶圓；每條邊補中點，邊才維持直的。"""
        S = self._S
        rx = r if r is not None else min((x1 - x0) * 0.22, (y1 - y0) * 0.22)
        mx, my = (x0 + x1) / 2, (y0 + y1) / 2
        pts = [x0 + rx, y0, mx, y0, x1 - rx, y0,
               x1, y0 + rx, x1, my, x1, y1 - rx,
               x1 - rx, y1, mx, y1, x0 + rx, y1,
               x0, y1 - rx, x0, my, x0, y0 + rx]
        self.cv.create_polygon([S(v) for v in pts], fill=fill,
                               outline=ink or self.INK,
                               width=max(1, S(w)), smooth=True)

    def _paw_row(self, x0, gap, y, n, col, wdt=5.6, hgt=7.5, stride=0.0):
        """腳。走路時前後腳一前一後錯開，站著時 stride 給 0 就是原本那排。"""
        S = self._S
        for i in range(n):
            paw = x0 + i * gap + (stride * 1.6 if i % 2 else -stride * 1.6)
            self.cv.create_oval(S(paw - wdt / 2), S(y - hgt),
                                S(paw + wdt / 2), S(y), fill=col,
                                outline=self.INK, width=max(1, S(1.8)))

    def _doodle_cat(self, x, t, y, stride=0.0, pose=None):
        S = self._S
        sway = math.sin(t * .04) * 2.0
        sink = 5 if pose in ("pee", "poop") else 0
        nod = 5 if pose == "eat" else 0
        self.cv.create_oval(S(x - 15), S(y - 2), S(x + 16), S(y + 3),
                         fill=shade(self._floor, -0.10), outline="")
        self.cv.create_line(S(x + 11), S(y - 12), S(x + 17), S(y - 16 + sway),
                         S(x + 19), S(y - 24 + sway),
                         fill=self.INK, width=max(2, S(5.6)), smooth=True,
                         capstyle="round")
        self.cv.create_line(S(x + 11), S(y - 12), S(x + 17), S(y - 16 + sway),
                         S(x + 19), S(y - 24 + sway),
                         fill=self.CAT_COAT, width=max(1, S(3.0)),
                         smooth=True, capstyle="round")
        self._paw_row(x - 6, 7.5, y, 3, self.CAT_COAT, stride=stride)
        self._blob(x - 10, y - 18 + sink, x + 12, y - 4, self.CAT_COAT,
                   2.1, r=5)
        for i in range(2):
            self.cv.create_line(S(x - 1 + i * 6), S(y - 16),
                             S(x - 2 + i * 6), S(y - 10),
                             fill=self.CAT_MARK, width=max(1, S(2.0)))
        hy = y - 24 + sink + nod
        for sd in (-1, 1):
            self.cv.create_polygon(S(x - 9 + (0 if sd < 0 else 12)), S(hy - 5),
                                S(x - 6 + (0 if sd < 0 else 12)), S(hy - 12),
                                S(x - 2 + (0 if sd < 0 else 12)), S(hy - 4),
                                fill=self.CAT_COAT, outline=self.INK,
                                width=max(1, S(1.9)))
        self._blob(x - 11, hy - 6, x + 5, hy + 7, self.CAT_COAT, 2.1, r=4.5)
        for sd in (-1, 1):
            self.cv.create_oval(S(x - 3 + sd * 3.6 - 1.4), S(hy - 1),
                             S(x - 3 + sd * 3.6 + 1.4), S(hy + 1.8),
                             fill=self.INK, outline="")
            for k in (-1.2, 0.6):
                self.cv.create_line(S(x - 3 + sd * 5), S(hy + 3 + k),
                                 S(x - 3 + sd * 12), S(hy + 2 + k * 1.6),
                                 fill=self.INK, width=max(1, S(1.1)))
        self.cv.create_polygon(S(x - 4.4), S(hy + 2.6), S(x - 1.6), S(hy + 2.6),
                            S(x - 3), S(hy + 4.4), fill="#c98a86", outline="")

    def _doodle_dog(self, x, t, y, stride=0.0, face=1, pose=None):
        S = self._S
        wag = math.sin(t * .16) * 3.2
        sink = 6 if pose in ("pee", "poop") else 0
        nod = 6 if pose == "eat" else 0
        self.cv.create_oval(S(x - 16), S(y - 2), S(x + 17), S(y + 3),
                         fill=shade(self._floor, -0.10), outline="")
        self.cv.create_line(S(x + 12 * face), S(y - 14), S(x + 18 * face),
                         S(y - 20), S(x + 17 * face + wag), S(y - 27),
                         fill=self.INK, width=max(2, S(6.4)), smooth=True,
                         capstyle="round")
        self.cv.create_line(S(x + 12 * face), S(y - 14), S(x + 18 * face),
                         S(y - 20), S(x + 17 * face + wag), S(y - 27),
                         fill=self.DOG_MARK, width=max(1, S(3.6)),
                         smooth=True, capstyle="round")
        self._paw_row(x - 7, 8, y, 3, self.DOG_COAT, 4.6, 6,
                      stride=stride)
        self._blob(x - 11, y - 19 + sink, x + 13, y - 5, self.DOG_COAT,
                   2.2, r=5.5)
        self._blob(x + 2, y - 18 + sink, x + 12, y - 7, self.DOG_MARK,
                   1.6, r=3.5)
        hy = y - 26 + sink + nod
        hx = x - 10 * face
        self._blob(hx - 9.5, hy - 8, hx + 9.5, hy + 9, self.DOG_COAT, 2.2,
                   r=5.5)
        self._blob(hx - 13 * face, hy - 7, hx - 5 * face, hy + 7,
                   self.DOG_MARK, 2.0, r=3.4)
        self._blob(hx + 1.5, hy - 8, hx + 9, hy + 1, self.DOG_MARK, 1.6, r=3)
        for sd in (-1, 1):
            self.cv.create_oval(S(hx + sd * 4 - 1.6), S(hy - 3.6),
                             S(hx + sd * 4 + 1.6), S(hy - 0.6),
                             fill=self.INK, outline="")
        self.cv.create_oval(S(hx - 5), S(hy + 1.6), S(hx + 2.4), S(hy + 8),
                         fill="#f8f4ec", outline=self.INK,
                         width=max(1, S(1.8)))
        self.cv.create_oval(S(hx - 2.6), S(hy + 2.2), S(hx + 0.6), S(hy + 4.6),
                         fill=self.INK, outline="")
        self.cv.create_oval(S(hx - 2.2), S(hy + 6.4), S(hx + 1.4), S(hy + 10),
                         fill="#c9564f", outline=self.INK,
                         width=max(1, S(1.5)))


class SketchPet(tk.Canvas):
    """一隻站著的貓或狗，放在標題旁邊當圖示。尾巴會動，人不動。"""

    UNIT_W, UNIT_H, CX, FLR = 56.0, 44.0, 28.0, 39.0

    def __init__(self, parent, kind="cat", size=84, page="#f4ecdc",
                 line="#a98c63", animate=True, **ignored):
        self.kind = kind
        self._page = page
        self.k = size / self.UNIT_H
        self._frame = 0
        self._alive = True
        super().__init__(parent, width=math.ceil(self.UNIT_W * self.k) + 1,
                         height=math.ceil(size) + 1, bg=page,
                         highlightthickness=0, bd=0)
        self.painter = PetPainter(self, self.k, line)
        tk.Canvas.bind(self, "<Configure>", lambda _e: self._redraw())
        tk.Canvas.bind(self, "<Destroy>", self._on_destroy)
        self._redraw()
        if animate:
            self.after(160, self._tick)

    def _on_destroy(self, _event=None):
        self._alive = False

    def _tick(self):
        if not self._alive:
            return
        try:
            if self.winfo_ismapped():
                self._frame += 1
                self._redraw()
            self.after(160, self._tick)
        except tk.TclError:
            self._alive = False

    def retheme(self, page=None, line=None, **ignored):
        self._page = page or self._page
        self.painter.line = line or self.painter.line
        try:
            tk.Canvas.configure(self, bg=self._page)
        except tk.TclError:
            return
        self._redraw()

    def _redraw(self):
        self.delete("all")
        self.painter.draw(self.kind, self.CX, self.FLR, self._frame)


class PetLane(tk.Canvas):
    """頁尾那條走道：貓和狗在裡面走來走去。

    走到邊界就轉頭，轉頭時整隻鏡射，臉一定朝著行進方向 —— 圖是朝左畫的，
    往右走不鏡射就是倒退嚕。到邊界會先停一下再回頭，不是撞牆彈回去。
    """

    UNIT_H, FLR = 44.0, 39.0
    EDGE = 29.0

    def __init__(self, parent, width=340, height=90, page="#f4ecdc",
                 line="#a98c63", **ignored):
        self._page = page
        self.k = height / self.UNIT_H
        self._alive = True
        super().__init__(parent, width=int(width), height=int(height),
                         bg=page, highlightthickness=0, bd=0)
        self.painter = PetPainter(self, self.k, line)
        span = width / self.k
        self.actors = [
            dict(kind="dog", x=span * 0.30, d=1, speed=0.62, wait=0,
                 frame=0),
            dict(kind="cat", x=span * 0.70, d=-1, speed=0.45, wait=18,
                 frame=40),
        ]
        tk.Canvas.bind(self, "<Destroy>", self._on_destroy)
        self._redraw()
        self.after(90, self._tick)

    def _on_destroy(self, _event=None):
        self._alive = False

    def retheme(self, page=None, line=None, **ignored):
        self._page = page or self._page
        self.painter.line = line or self.painter.line
        try:
            tk.Canvas.configure(self, bg=self._page)
        except tk.TclError:
            return
        self._redraw()

    def _bounds(self):
        span = (self.winfo_width() or int(self["width"])) / self.k
        return self.EDGE, max(self.EDGE + 4.0, span - self.EDGE)

    MIN_GAP = 36.0

    def _step(self):
        left, right = self._bounds()
        for pet in self.actors:
            pet["frame"] += 1
            if pet["wait"] > 0:
                pet["wait"] -= 1
                continue
            pet["x"] += pet["d"] * pet["speed"]
            if pet["x"] <= left:
                pet["x"], pet["d"], pet["wait"] = left, 1, 12
            elif pet["x"] >= right:
                pet["x"], pet["d"], pet["wait"] = right, -1, 12

        first, second = self.actors[0], self.actors[1]
        if abs(first["x"] - second["x"]) < self.MIN_GAP:
            near, far = ((first, second) if first["x"] <= second["x"]
                         else (second, first))
            near["d"], far["d"] = -1, 1
            near["wait"] = far["wait"] = 14
            overlap = self.MIN_GAP - (far["x"] - near["x"])
            near["x"] = max(left, near["x"] - overlap / 2)
            far["x"] = min(right, far["x"] + overlap / 2)

    def _tick(self):
        if not self._alive:
            return
        try:
            if self.winfo_ismapped():
                self._step()
                self._redraw()
            self.after(90, self._tick)
        except tk.TclError:
            self._alive = False

    def _redraw(self):
        self.delete("all")
        for pet in self.actors:
            self.painter.draw(pet["kind"], pet["x"], self.FLR, pet["frame"],
                              facing=pet["d"], walking=pet["wait"] == 0)


class SketchRule(tk.Canvas):
    """標題底下那一條線。原本是一整排 ══，改成一筆畫過去的波浪線。

    三件事讓它像筆畫而不像函數圖形：兩端收細（起筆收筆有壓力變化）、
    起伏中間大兩端小（手腕擺動的樣子）、旁邊兩道蠟筆色的短筆錯開一點。
    """

    def __init__(self, parent, width=260, page="#fdf4e6", ink="#d98f4f",
                 **ignored):
        self._page, self._ink = page, ink
        super().__init__(parent, width=width, height=px(26), bg=page,
                         highlightthickness=0, bd=0)
        tk.Canvas.bind(self, "<Configure>", lambda _e: self._redraw())

    def configure(self, **kwargs):
        for key in ("bg", "background"):
            if key in kwargs:
                self._page = kwargs.pop(key)
        for key in ("fg", "foreground"):
            if key in kwargs:
                self._ink = kwargs.pop(key)
        if kwargs:
            tk.Canvas.configure(self, **kwargs)
        tk.Canvas.configure(self, bg=self._page)
        self._redraw()

    config = configure

    def _pen(self, x0, x1, mid, amp, drop, cycles, seed, colour, thickness):
        """畫一道筆。分段畫是為了讓線寬沿著筆畫變化 —— 一整條同寬的線
        是印出來的，不是畫出來的。"""
        steps = 44
        points = []
        for step in range(steps):
            ratio = step / (steps - 1.0)
            envelope = math.sin(math.pi * ratio) ** 0.55
            points.append((
                x0 + (x1 - x0) * ratio,
                mid + drop
                + math.sin(ratio * math.pi * cycles) * amp * envelope
                + _jitter(seed, step, 0.55)))
        for index in range(len(points) - 1):
            ratio = index / (len(points) - 2.0)
            taper = 0.3 + 0.7 * math.sin(math.pi * ratio) ** 0.7
            (ax, ay), (bx, by) = points[index], points[index + 1]
            self.create_line(ax, ay, bx, by, fill=colour,
                             width=max(1.0, thickness * taper),
                             capstyle="round")

    def _sparkle(self, x, y, size, colour, width=1.3):
        for dx, dy in ((0, -1), (0, 1), (-0.9, -0.45), (0.9, 0.45),
                       (-0.9, 0.45), (0.9, -0.45)):
            self.create_line(x, y, x + dx * size, y + dy * size,
                             fill=colour, width=width, capstyle="round")

    def _redraw(self):
        self.delete("all")
        width = self.winfo_width() or int(self["width"])
        height = self.winfo_height() or int(self["height"])
        if width <= 1:
            return
        mid = height / 2.0
        left, right = px(26), width - px(30)
        if right - left < px(40):
            return
        span = right - left
        self._pen(left, right, mid, px(4.4), 0, 2.4, 1.0,
                  self._ink, px(2.6))
        self._pen(left + span * 0.30, right - span * 0.26, mid,
                  px(3.0), px(2.6), 1.6, 6.1, POP["mint"], px(1.5))
        self._pen(left + span * 0.12, right - span * 0.50, mid,
                  px(2.6), -px(2.6), 1.4, 2.7, POP["sky"], px(1.5))
        self._sparkle(width - px(16), mid - px(6), px(5), POP["coral"],
                      max(1.0, px(1.3)))
        self._sparkle(px(15), mid + px(5), px(3), POP["sun"],
                      max(1.0, px(1.2)))


def sketch_oval_path(cx, cy, rx, ry, seed=0.0, wobble=0.8, steps=16):
    """畫歪的橢圓。貓狗的頭、耳朵、腳掌都是這個。"""
    points = []
    for index in range(steps + 1):
        angle = math.radians(360.0 * index / steps)
        points.append(cx + rx * math.cos(angle) + _jitter(seed, index, wobble))
        points.append(cy + ry * math.sin(angle)
                      + _jitter(seed + 3.3, index, wobble))
    return points


def _tree_font():
    """清單實際在用的字體。欄寬要拿它來量 —— 量 TkDefaultFont 會短一截，
    字就被切在一半。"""
    try:
        return tkfont.Font(font=data_font(10))
    except Exception:
        return tkfont.nametofont("TkDefaultFont")


def apply_ui_fonts():
    """把 Tk 內建的那幾支具名字體換成內文那一隻。

    直接寫在 widget 上的字體只管得到自己那一顆；對話框、選單、提示用的是
    這幾支具名字體 —— 不換，畫面上就會一半是選的字體、一半是系統預設。
    Windows 的 messagebox 是作業系統自己畫的，換不到。
    """
    for name in ("TkDefaultFont", "TkTextFont", "TkMenuFont", "TkHeadingFont",
                 "TkCaptionFont", "TkSmallCaptionFont", "TkIconFont",
                 "TkTooltipFont"):
        try:
            tkfont.nametofont(name).configure(family=UI_FONT)
        except (tk.TclError, RuntimeError, ValueError):
            pass
    try:
        tkfont.nametofont("TkFixedFont").configure(family=UI_FONT_DATA)
    except (tk.TclError, RuntimeError, ValueError):
        pass

def scrollable_page(parent, theme_bg="#D6C0A9", bar_parent=None):
    """A page that grows a vertical scrollbar when the window is too short.

    Returns (outer, inner, canvas, scrollbar). Build into `inner`; pack
    `outer` where the page used to go.

    A tab taller than its window simply hides its own bottom edge, with no
    cue that anything is missing -- so a button below the fold does not
    exist as far as the person using it is concerned.

    bar_parent -- where the scrollbar itself lives. Left out, it sits inside
    the page and is therefore exactly as tall as the scrolling area. Pass the
    tab when something is pinned below the page (the yard, a footer): the bar
    then runs the full height of the window instead of stopping short of the
    bottom with a stub of bare background under it. Pass the tab BEFORE
    anything else is packed into it, so the bar claims its full-height strip
    down the right edge first and everything after it packs to the left.
    """
    outer = tk.Frame(parent, bg=theme_bg)
    canvas = tk.Canvas(outer, bg=theme_bg, highlightthickness=0, bd=0)
    bar_host = outer if bar_parent is None else bar_parent
    bar = SketchScrollbar(bar_host, orient="vertical", command=canvas.yview,
                          page=theme_bg)
    inner = tk.Frame(canvas, bg=theme_bg)

    window = canvas.create_window((0, 0), window=inner, anchor="nw")
    canvas.configure(yscrollcommand=bar.set)

    def _resize(_event=None):
        canvas.configure(scrollregion=canvas.bbox("all"))
        width = canvas.winfo_width()
        if width > 1:
            canvas.itemconfigure(window, width=width)

    inner.bind("<Configure>", _resize)
    canvas.bind("<Configure>", _resize)

    def _wheel(event):
        number = getattr(event, "num", None)
        step = -1 if number == 4 else (
            1 if number == 5 else
            int(-1 * (event.delta / 120)) if getattr(event, "delta", 0) else 0)
        if step:
            canvas.yview_scroll(step, "units")

    def _bind_wheel(_event=None):
        for sequence in ("<MouseWheel>", "<Button-4>", "<Button-5>"):
            canvas.bind_all(sequence, _wheel)

    def _unbind_wheel(_event=None):
        for sequence in ("<MouseWheel>", "<Button-4>", "<Button-5>"):
            canvas.unbind_all(sequence)

    canvas.bind("<Enter>", _bind_wheel)
    canvas.bind("<Leave>", _unbind_wheel)

    canvas.pack(side="left", fill="both", expand=True)
    bar.pack(side="right", fill="y")
    return outer, inner, canvas, bar


THEME_FILE = Path.home() / ".warehouse_billing_theme"
THEME_FILE_VERSION = "v1"


def palette(name):
    """把 Audit Center 的角色名對到這支檔案原本用的短名，舊的呼叫端不用動。"""
    t = UI_THEMES[name if name in UI_THEMES else "Crayon"]
    out = dict(t)
    out.update(bg=t["BG"], side=t["PANEL"], line=t["LINE"], ink=t["TEXT"],
               dim=t["ACCENT"], accent=t["TITLE"], hover=t["RUN_HOVER"],
               sel=t["SELECT"], zebra=t["STRIPE"], run=t["RUN"],
               tab=t["TAB_BG"])
    return out


C = palette("Crayon")


CHARGES = [
 ("c20",   "inbound", "卸櫃分類 20 呎櫃", "Unload and sort, 20GP", "櫃", "container", 380.0, "20GP"),
 ("c40",   "inbound", "卸櫃分類 40 呎 / 40 高 / 45 呎櫃", "Unload and sort, 40GP / 40HQ / 45GP", "櫃", "container", 550.0, "40GP/40HQ/45GP"),
 ("c53",   "inbound", "卸櫃分類 53 呎乾櫃", "Unload and sort, 53 dry van", "櫃", "container", 580.0, "53 DRY VAN"),
 ("blind", "inbound", "未預約收貨", "Blind receiving", "櫃", "container", 50.0, "Blind receiving"),
 ("yard",  "inbound", "空櫃滯留費", "Yard fee", "櫃天", "container-day", 50.0, "Yard fee"),
 ("prcv",  "inbound", "棧板收貨", "Pallet receive", "板", "pallet", 15.0, "Pallet receive"),
 ("pin",   "inbound", "進倉打板", "Palletizing on receipt", "板", "pallet", 6.0, "Palletizing (1 Sku)"),
 ("sort",  "inbound", "分類費 超過 8 個品項", "Sorting charge beyond 8 SKUs", "品項", "SKU", 8.0, "Sorting Charge(per SKU)"),
 ("lblin", "inbound", "貼標", "Labelling", "張", "label", 0.35, "Labelling"),
 ("i1",  "inbound", "收貨 未滿 1 磅", "Receiving under 1 lb", "件", "piece", 0.30, "<1.00 lbs"),
 ("i5",  "inbound", "收貨 1 至 5 磅", "Receiving 1 to 5 lbs", "件", "piece", 0.55, "Box >1.00 - 5.00 lbs"),
 ("i10", "inbound", "收貨 5 至 10 磅", "Receiving 5 to 10 lbs", "件", "piece", 1.00, "Box >5.00 - 10.00 lbs"),
 ("i20", "inbound", "收貨 10 至 20 磅", "Receiving 10 to 20 lbs", "件", "piece", 1.50, "Box >10.00 - 20.00 lbs"),
 ("i40", "inbound", "收貨 20 至 40 磅", "Receiving 20 to 40 lbs", "件", "piece", 2.50, "Box >20.00 - 40.00 lbs"),
 ("i50", "inbound", "收貨 40 至 50 磅", "Receiving 40 to 50 lbs", "件", "piece", 2.75, "Box >40.00 - 50.00 lbs"),
 ("i70", "inbound", "收貨 50 至 70 磅", "Receiving 50 to 70 lbs", "件", "piece", 3.00, "Box >50.00 - 70.00 lbs"),
 ("i99", "inbound", "收貨 70 磅以上", "Receiving over 70 lbs", "件", "piece", 3.40, "Box 70.00 + lbs"),

 ("lblout", "outbound", "出貨貼標", "Outbound label", "張", "label", 0.20, "Label"),
 ("o1",  "outbound", "逐件揀貨 未滿 1 磅", "Piece pick, under 1 lb", "件", "piece", 0.38, "<1.00 lbs"),
 ("o5",  "outbound", "逐件揀貨 1 至 5 磅", "Piece pick, 1 to 5 lbs", "件", "piece", 1.08, ">1.00 - 5.00 lbs"),
 ("o10", "outbound", "逐件揀貨 5 至 10 磅", "Piece pick, 5 to 10 lbs", "件", "piece", 1.34, ">5.00 - 10.00 lbs"),
 ("o20", "outbound", "逐件揀貨 10 至 20 磅", "Piece pick, 10 to 20 lbs", "件", "piece", 1.64, ">10.00 - 20.00 lbs"),
 ("o40", "outbound", "逐件揀貨 20 至 40 磅", "Piece pick, 20 to 40 lbs", "件", "piece", 2.60, ">20.00 - 40.00 lbs"),
 ("o50", "outbound", "逐件揀貨 40 至 50 磅 兩人作業", "Piece pick, 40 to 50 lbs, two people", "件", "piece", 2.85, ">40.00 - 50.00 lbs"),
 ("o70", "outbound", "逐件揀貨 50 至 70 磅 兩人作業", "Piece pick, 50 to 70 lbs, two people", "件", "piece", 3.25, ">50.00 - 70.00 lbs"),
 ("o99", "outbound", "逐件揀貨 70 磅以上 兩人作業", "Piece pick, over 70 lbs, two people", "件", "piece", 4.00, "70.00 + lbs"),
 ("ship", "outbound", "出貨標籤", "Shipping label", "張", "label", 3.00, "Shipping Label"),
 ("bol",  "outbound", "貨運提單", "Trucking bill of lading", "趟", "shipment", 10.0, "Trucking BOL"),
 ("ppick","outbound", "整板揀貨", "Full pallet pick", "板", "pallet", 15.0, "Pallet Pick(1SKU)"),
 ("cont", "outbound", "代客聯繫", "Contact service", "次", "request", 5.0, "Contact Service"),
 ("mord", "outbound", "人工建單", "Manual order processing", "單", "order", 3.0, "Manual Order Processing"),
 ("pout", "outbound", "出貨打板", "Palletizing on despatch", "板", "pallet", 15.0, "Palletizing"),

 ("stor",  "storage", "一般倉租", "General storage", "立方米天", "CBM-day", 0.40, "General Storage"),
 ("rk1",   "storage", "貨架附加費 小型架", "Racking surcharge, small rack", "月", "month", 27.50, None),
 ("rk8",   "storage", "貨架附加費 8 呎架位", "Racking surcharge, 8 foot section", "月", "month", 32.50, None),
 ("rk12",  "storage", "貨架附加費 12 呎架位", "Racking surcharge, 12 foot section", "月", "month", 42.50, None),
 ("rk12d", "storage", "貨架附加費 12 呎雙深架位", "Racking surcharge, 12 foot double deep", "月", "month", 80.00, None),
 ("labor", "storage", "人工 半小時計", "Hourly labor, half hour blocks", "小時", "hour", 33.0, "Hourly Labor"),
 ("ot",    "storage", "加班人工 超過 8 小時", "Overtime labor beyond 8 hours", "小時", "hour", 49.50, "Overtime"),
 ("wrap",  "storage", "膠膜 出貨棧板免費", "Stretch wrap, waived on outbound", "次", "job", None, "Stretch Wrap"),
 ("sup",   "storage", "耗材加成 10%", "Supplies, 10% mark-up", "筆", "item", None, "Supplies"),
 ("pltr",  "storage", "出貨棧板 一般 B 級", "Outbound pallet, grade B regular", "板", "pallet", 20.0, None),
 ("pltl",  "storage", "出貨棧板 大型 B 級", "Outbound pallet, grade B large", "板", "pallet", 30.0, None),

 ("rpic",  "return", "加拍照片", "Additional pictures", "張", "picture", 0.75, "Pictures"),
 ("r1",    "return", "退貨 未滿 1 磅 含 2 張照片與貼標", "Return under 1 lb, 2 pictures", "件", "piece", 1.15, "<1.00 lbs"),
 ("r20",   "return", "退貨 1 至 20 磅 含 3 張照片與貼標", "Return 1 to 20 lbs, 3 pictures", "件", "piece", 2.25, ">1.00-20.00 lbs"),
 ("r40",   "return", "退貨 20 至 40 磅 含 3 張照片與貼標", "Return 20 to 40 lbs, 3 pictures", "件", "piece", 3.00, ">20.00-40.00 lbs"),
 ("r99",   "return", "退貨 40 磅以上 含 3 張照片與貼標", "Return over 40 lbs, 3 pictures", "件", "piece", 4.00, "40.00 + lbs"),
 ("rplt",  "return", "退貨整板揀貨", "Return pallet pick", "板", "pallet", 15.0, "Pallet Pick(1SKU)"),
 ("repack","return", "重新包裝", "Repack", "件", "piece", None, "Repack"),
 ("rsup",  "return", "退貨耗材加成 10%", "Return supplies, 10% mark-up", "筆", "item", None, "Supplies"),
 ("pmat",  "return", "包裝材料", "Packing material", "筆", "item", None, "Packing Material"),
 ("rstor", "return", "退貨倉租", "Return storage", "板", "pallet", 20.0, "Return Storage"),
 ("disp",  "return", "銷毀處理", "Disposal", "筆", "item", None, "Disposal"),
]
KEYS = [c[0] for c in CHARGES]
CH = {c[0]: dict(zip(("grp", "zh", "en", "u_zh", "u_en", "rate", "sheet"),
                     c[1:])) for c in CHARGES}
GROUPS = ["inbound", "outbound", "storage", "return"]

T = {
"zh": {
 "app": "倉儲計費", "sub": "建層級 · 設費率 · 填數量 · 看合計",
 "clients": "客戶", "add": "＋ 新增客戶", "del": "刪除客戶",
 "code": "代碼", "name": "名稱", "colour": "配色",
 "pages": ("① 層級", "② 客戶", "③ 費率", "④ 計算", "⑤ 帳單"),
 "client": "客戶", "contact": "聯絡人", "note": "備註",
 "cl_new": "新增客戶", "cl_edit": "修改選定客戶", "cl_list": "客戶清單",
 "money": "${v}",
 "levels": "費率層級", "level": "費率層級", "editing": "正在編輯",
 "find_client": "搜尋客戶", "find_level": "搜尋層級",
 "find_hint": "輸入代碼、名稱、聯絡人或層級過濾，點一列就帶進這一頁",
 "pick_cols": ("代碼", "名稱", "費率層級"),
 "p1_title": "費率層級", "p2_title": "客戶", "p3_title": "費率",
 "p4_title": "本期計算", "p5_title": "帳單",
 "lv_name": "層級名稱", "lv_new": "＋ 新增層級", "lv_copy": "複製選定層級",
 "lv_rename": "改名", "lv_del": "刪除層級", "lv_list": "層級清單",
 "lv_cols": ("層級", "客戶數", "已設單價", "狀態"),
 "lv_default": "預設", "join": "、",
 "lv_blank": "層級要有名稱。", "lv_dup": "已經有一層叫這個名字。",
 "lv_last": "至少要留一層。",
 "lv_none": "還沒有費率層級。到「① 層級」建一層。",
 "lv_inuse": "「{name}」還有 {n} 個客戶在用，不能刪。先到「② 客戶」把他們改派到別的層級：{who}",
 "lv_del_ask": "刪除「{name}」？那一層的整份單價會一起消失，沒有復原。",
 "lv_added": "已新增「{name}」，單價從報價表預設值起頭。到「③ 費率」改。",
 "lv_copied": "已複製成「{name}」，現在編的是新的那一層。",
 "lv_users": "{n} 個客戶用這一層", "lv_nobody": "還沒有客戶指到這一層",
 "rate_for": "正在編輯：{name}", "lv_of": "費率層級：{name}",
 "unset": "未指定", "gone": "{id}（層級已刪除）",
 "no_level": "這個客戶還沒指定費率層級，算不出金額。到「② 客戶」指定一層。",
 "lost_level": "這個客戶指到的層級「{id}」已經不在了。到「② 客戶」重新指定。",
 "need_level": "先到「② 客戶」替這個客戶指定費率層級。",
 "p1_help": "先建費率層級。一層是一整份獨立的單價表，好幾個客戶可以共用同一層；點一列就切成正在編輯的那一層。",
 "p2_help": "再建客戶，並指定它用哪一層費率。代碼是唯一的，每期的數量掛在代碼下；單價則跟著層級走。",
 "p3_help": "設定正在編輯那一層的單價。上方選分類，只顯示那一類的項目。改一次，指到這一層的客戶全部跟著變。",
 "p4_help": "把這期做的量一項一項加進來：選項目、填數量、按「加入」。單價從這個客戶的層級帶入。",
 "p5_help": "只列出有數量的項目，可以匯出成帳單。",
 "need_client": "還沒有客戶。到「② 客戶」新增一個。",
 "group": "分類", "all": "全部",
 "item": "項目", "qty": "數量", "unit": "單位", "price": "單價",
 "amount": "金額", "add": "加入", "rm": "刪除選定",
 "added": "已加入的項目", "calc_cols": ("單號", "項目", "數量", "單位", "單價", "金額"),
 "pick_card": "出入庫作業快速輸入", "order_no": "單號", "flow_type": "作業類型",
 "inbound_flow": "入庫", "outbound_flow": "出庫", "pick_mode": "計費方式",
 "receive_piece": "逐件入庫", "piece_pick": "逐件揀貨",
 "work_date": "作業日期", "weight_each": "單件重量（lb）", "container_qty": "櫃數", "sku_count": "SKU 數",
 "add_pick": "加入作業費", "edit_selected": "修改選定", "save_edit": "儲存修改", "cancel_edit": "取消修改", "pick_hint": "先選入庫或出庫，再選下一層計費方式。入庫可另外填櫃數與 SKU 數，逐件作業會依單件重量自動帶入費率級距。",
 "need_order": "請輸入單號。", "need_weight": "重量要填大於 0 的數字。",
 "need_pick_rate": "此揀貨方式尚未設定單價，請先到「③ 費率」頁設定。",
 "storage_card": "倉儲費快速計算", "storage_hint": "選擇計費期間並輸入平均占用 CBM，系統會自動計算天數與 CBM-day。單價可在「③ 費率」頁的倉租分類修改。",
 "storage_start": "起始日期", "storage_end": "結束日期", "storage_cbm": "平均占用（CBM）", "storage_days": "計費天數", "storage_rate": "單價 / CBM-day", "storage_amount": "預估金額", "add_storage": "加入倉儲費", "need_storage_date": "請輸入有效的倉儲起訖日期。", "need_storage_cbm": "平均占用 CBM 必須大於 0。",
 "pick_item": "請先選一個項目。", "need_qty": "數量要填大於 0 的數字。",
 "need_amt": "這個項目報價未定，請直接填金額。",
 "cl_cols": ("代碼", "名稱", "費率層級", "聯絡人", "備註"),
 "bill_title": "帳單", "bill_sub": "只列出有數量的項目。",
 "bill_cols": ("單號", "項目", "數量", "單位", "單價", "金額"),
 "bill_empty": "本期還沒有填任何數量。到「④ 計算」填數量。",
 "exp_bill": "匯出帳單", "bill_tpl": "帳單模板",
 "grp": {"inbound": "進倉", "outbound": "出貨",
         "storage": "倉租", "return": "退貨"},
 "subtotal": "{g}小計", "byquote": "報價未定", "typed": "自填",
 "period": "本期合計", "clear": "清空本期",
 "clear_ask": "本期填的數量會全部清掉，確定嗎？",
 "save": "存檔", "saved": "已存檔", "imp": "匯入報價表", "exp": "匯出報價表模板",
 "warn": "請檢查", "ok": "完成", "export_ok": "匯出完成\n\n檔案已儲存到：\n{path}", "no_bill_data": "目前沒有任何可匯出的帳單明細。請先在「④ 計算」頁加入入庫、出庫或其他收費項目。",
 "empty": "還沒有客戶。",
 "b_title": "先建一層費率，再新增客戶",
 "b_body": "「① 層級」建一層，例如  A 級；「② 客戶」填代碼和名稱，例如  GEN  和  Geniqua Client，\n"
           "並指到那一層。之後：③ 設那一層的單價　④ 填數量算錢　⑤ 匯出帳單。\n"
           "也可以按右上角「匯入報價表」直接把 Excel 讀進正在編輯的層級。",
 "blank": "代碼和名稱都要填。", "dup": "代碼重複。",
 "imp_ok": "已從報價表更新「{lv}」的 {n} 個項目。",
 "imp_fail": "讀不到費率。請確認是報價表格式（Price 工作表）。",
 "need_xl": "匯入 Excel 需要 openpyxl：pip install openpyxl",
},
"en": {
 "app": "Warehouse Billing", "sub": "levels · rates · quantities · total",
 "clients": "Clients", "add": "+ New client", "del": "Delete client",
 "code": "Code", "name": "Name", "colour": "Colour",
 "pages": ("1  Levels", "2  Clients", "3  Rates", "4  Calculate", "5  Invoice"),
 "client": "Client", "contact": "Contact", "note": "Note",
 "cl_new": "Add client", "cl_edit": "Update selected", "cl_list": "Client list",
 "money": "${v}",
 "levels": "Rate levels", "level": "Rate level", "editing": "Editing",
 "find_client": "Find client", "find_level": "Find level",
 "find_hint": "Type a code, name, contact or level to filter; click a row to bring it into this page",
 "pick_cols": ("Code", "Name", "Rate level"),
 "p1_title": "Rate levels", "p2_title": "Clients", "p3_title": "Rates",
 "p4_title": "This period", "p5_title": "Invoice",
 "lv_name": "Level name", "lv_new": "+ New level", "lv_copy": "Duplicate selected",
 "lv_rename": "Rename", "lv_del": "Delete level", "lv_list": "Level list",
 "lv_cols": ("Level", "Clients", "Rates set", "State"),
 "lv_default": "Default", "join": ", ",
 "lv_blank": "Give the level a name.", "lv_dup": "There is already a level with that name.",
 "lv_last": "Keep at least one level.",
 "lv_none": "No rate levels yet. Add one on the Levels page.",
 "lv_inuse": "\"{name}\" is still assigned to {n} client(s) and cannot be deleted. Reassign them on the Clients page first: {who}",
 "lv_del_ask": "Delete \"{name}\"? That level's entire rate set goes with it, and there is no undo.",
 "lv_added": "Added \"{name}\" with the price-sheet defaults. Change them on the Rates page.",
 "lv_copied": "Copied to \"{name}\" — you are now editing the copy.",
 "lv_users": "{n} client(s) on this level", "lv_nobody": "No client is on this level yet",
 "rate_for": "Editing: {name}", "lv_of": "Rate level: {name}",
 "unset": "Not set", "gone": "{id} (level deleted)",
 "no_level": "This client has no rate level, so nothing can be priced. Assign one on the Clients page.",
 "lost_level": "This client's level \"{id}\" no longer exists. Reassign one on the Clients page.",
 "need_level": "Assign this client a rate level on the Clients page first.",
 "p1_help": "Create rate levels first. A level is one complete, independent set of rates; any number of clients can share it. Click a row to make it the level being edited.",
 "p2_help": "Then add clients and assign each one a level. The code is unique; each period's quantities hang off it, while the rates follow the level.",
 "p3_help": "Set the rates of the level being edited. Choose a group above to show just those lines. Change a rate once and every client on this level follows.",
 "p4_help": "Add this period's work one line at a time: pick an item, enter a quantity, press Add. Rates come from this client's level.",
 "p5_help": "Only lines with a quantity. Export the invoice when you are done.",
 "need_client": "No clients yet. Add one on the Clients page.",
 "group": "Group", "all": "All",
 "item": "Item", "qty": "Qty", "unit": "Unit", "price": "Rate",
 "amount": "Amount", "add": "Add", "rm": "Remove selected",
 "added": "Added this period", "calc_cols": ("Order No.", "Item", "Qty", "Unit", "Rate", "Amount"),
 "pick_card": "Inbound / Outbound Quick Entry", "order_no": "Order No.", "flow_type": "Operation",
 "inbound_flow": "Inbound", "outbound_flow": "Outbound", "pick_mode": "Billing Method",
 "receive_piece": "Piece Receiving", "piece_pick": "Piece Pick",
 "work_date": "Activity Date", "weight_each": "Weight per Unit (lb)", "container_qty": "Containers", "sku_count": "SKU Count",
 "add_pick": "Add Activity Fee", "edit_selected": "Edit Selected", "save_edit": "Save Changes", "cancel_edit": "Cancel Edit", "pick_hint": "Choose Inbound or Outbound first, then choose the billing method. For inbound work, you may also enter the container count and SKU count. Piece-based work uses the weight bracket automatically.",
 "need_order": "Enter an order number.", "need_weight": "Weight must be a number above zero.",
 "need_pick_rate": "No rate is set for this pick type. Set it on the Rates page first.",
 "storage_card": "Storage Fee Calculator", "storage_hint": "Choose the billing period and enter the average occupied CBM. The tool calculates days and CBM-days automatically. Change the rate under Storage on the Rates page.",
 "storage_start": "Start Date", "storage_end": "End Date", "storage_cbm": "Average Occupancy (CBM)", "storage_days": "Billable Days", "storage_rate": "Rate / CBM-day", "storage_amount": "Estimated Amount", "add_storage": "Add Storage Fee", "need_storage_date": "Enter a valid storage start and end date.", "need_storage_cbm": "Average occupied CBM must be above zero.",
 "pick_item": "Choose an item first.", "need_qty": "Quantity must be a number above zero.",
 "need_amt": "This line has no set price -- enter the amount instead.",
 "cl_cols": ("Code", "Name", "Rate level", "Contact", "Note"),
 "bill_title": "Invoice", "bill_sub": "Only lines with a quantity.",
 "bill_cols": ("Order No.", "Item", "Qty", "Unit", "Rate", "Amount"),
 "bill_empty": "Nothing entered yet. Go to Calculate and enter quantities.",
 "exp_bill": "Export invoice", "bill_tpl": "Invoice template",
 "grp": {"inbound": "Inbound", "outbound": "Outbound",
         "storage": "Storage", "return": "Return"},
 "subtotal": "{g} subtotal", "byquote": "no set price", "typed": "typed",
 "period": "Period total", "clear": "Clear period",
 "clear_ask": "This wipes every quantity for this period. Continue?",
 "save": "Save", "saved": "Saved", "imp": "Import price sheet",
 "exp": "Export price sheet template",
 "warn": "Check this", "ok": "Done", "export_ok": "Export complete\n\nThe file was saved to:\n{path}", "no_bill_data": "There are no invoice details to export. Add inbound, outbound, or other charge lines on the Calculate page first.",
 "empty": "No clients yet.",
 "b_title": "Start with a rate level, then a client",
 "b_body": "Add a level on 1 Levels, for example  Level A.  Add a client on 2 Clients, for example\n"
           "GEN  and  Geniqua Client, and point it at that level. Then: 3 set the level's rates,\n"
           "4 enter quantities, 5 export the invoice. Import price sheet at the top right reads\n"
           "your Excel straight into the level being edited.",
 "blank": "Code and name are both required.", "dup": "That code is taken.",
 "imp_ok": "Updated {n} lines of \"{lv}\" from the price sheet.",
 "imp_fail": "No rates found. Expected the price sheet layout (Price worksheet).",
 "need_xl": "Importing Excel needs openpyxl:  pip install openpyxl",
},
}


class Book:
    """Rate levels, and per client this period's quantities.

    費率掛在「層級」上，不掛在客戶上。一層是一整份獨立的單價表；客戶指到
    哪一層，計算頁就用哪一層的單價。同一層可以給很多客戶共用 —— 改一次，
    指到那一層的客戶全部跟著變。這跟 UPS 對帳工具的費率層級是同一套做法。

    Quantities still live on the client rather than in an activity log
    because the calculation page IS the record -- fill the month's counts
    down the list and read the total. Nothing to log, nothing to undo."""

    def __init__(self):
        self.levels = {}    # id -> {"name", "rates": {key: float | None}}
        self.clients = {}   # code -> {"name", "contact", "note", "level",
                            #          "qty", "amt", "picks"}

    @staticmethod
    def fresh_rates():
        return {k: CH[k]["rate"] for k in KEYS}

    # ---- levels ----------------------------------------------------------

    def level_ids(self):
        """Creation order. The id is internal (L1, L2 ...); people only ever
        see the name, and the name can change."""
        return list(self.levels)

    def _next_level_id(self):
        n = 1
        while f"L{n}" in self.levels:
            n += 1
        return f"L{n}"

    def _check_level_name(self, name, skip=None):
        name = (name or "").strip()
        if not name:
            raise ValueError("lv_blank")
        for lid, lv in self.levels.items():
            if lid != skip and lv["name"].strip().lower() == name.lower():
                raise ValueError("lv_dup")
        return name

    def add_level(self, name, rates=None):
        """A new level starts from the published price sheet, so ③ is never
        a page of blanks -- the first job is to change the lines this level
        differs on. Pass rates to seed it from somewhere else."""
        name = self._check_level_name(name)
        lid = self._next_level_id()
        base = self.fresh_rates()
        if rates:
            base.update(rates)
        self.levels[lid] = {"name": name, "rates": base}
        return lid

    def copy_level(self, src, name):
        if src not in self.levels:
            raise ValueError("lv_none")
        return self.add_level(name, dict(self.levels[src]["rates"]))

    def rename_level(self, lid, name):
        if lid not in self.levels:
            raise ValueError("lv_none")
        self.levels[lid]["name"] = self._check_level_name(name, skip=lid)

    def level_users(self, lid):
        return sorted(c for c, d in self.clients.items()
                      if d.get("level") == lid)

    def delete_level(self, lid):
        """Refuse while any client points here. Deleting first and finding
        out at billing time that a whole batch has no prices is the wrong
        order -- say who has to be moved instead."""
        if lid not in self.levels:
            return
        if len(self.levels) < 2:
            raise ValueError("lv_last")
        if self.level_users(lid):
            raise ValueError("lv_inuse")
        del self.levels[lid]

    def level_name(self, lid):
        lv = self.levels.get(lid)
        return lv["name"] if lv else ""

    def ensure_level(self, name):
        """There is always at least one level, so ③ always has something to
        edit and a new client always has something to point at."""
        if not self.levels:
            self.add_level(name)
        return next(iter(self.levels))

    def level_rate(self, lid, key):
        return self.levels[lid]["rates"][key]

    def set_level_rate(self, lid, key, value):
        """None means the sheet has no set price -- Varies, a mark-up, or
        waived. Those get their amount typed in on the calculation page."""
        if value in (None, ""):
            self.levels[lid]["rates"][key] = None
            return
        v = float(value)
        if v < 0:
            raise ValueError("rate")
        self.levels[lid]["rates"][key] = v

    def rates_set(self, lid):
        return sum(1 for v in self.levels[lid]["rates"].values()
                   if v is not None)

    # ---- clients ---------------------------------------------------------

    def add_client(self, code, name, level=""):
        code, name = code.strip().upper(), name.strip()
        if not code or not name:
            raise ValueError("blank")
        if code in self.clients:
            raise ValueError("dup")
        self.clients[code] = {"name": name, "contact": "", "note": "",
                              "level": level if level in self.levels else "",
                              "qty": {}, "amt": {}, "picks": []}
        return code

    def update_client(self, code, name, contact, note, level=None):
        """level None leaves the assignment alone; "" clears it."""
        if code not in self.clients:
            raise ValueError("blank")
        name = name.strip()
        if not name:
            raise ValueError("blank")
        c = self.clients[code]
        c["name"], c["contact"], c["note"] = name, contact.strip(), note.strip()
        if level is not None:
            c["level"] = level

    def set_client_level(self, code, level):
        self.clients[code]["level"] = level or ""

    def delete_client(self, code):
        self.clients.pop(code, None)

    def resolve(self, code):
        """Which level prices this client: (level id, status). The two bad
        statuses are different things -- "none" was never assigned, "lost"
        points at a level that has since gone -- and both have to be said
        out loud. Showing a deleted level as "not set" makes it look like
        nobody ever chose one."""
        lid = self.clients[code].get("level") or ""
        if not lid:
            return "", "none"
        if lid not in self.levels:
            return lid, "lost"
        return lid, "ok"

    def priced(self, code):
        return self.resolve(code)[1] == "ok"

    def rate(self, code, key):
        """The client's rate for a line, through its level. None when the
        line has no set price -- or when the client has no usable level,
        which callers check first with priced()."""
        lid, status = self.resolve(code)
        if status != "ok":
            return None
        return self.levels[lid]["rates"][key]

    def qty(self, code, key):
        return self.clients[code]["qty"].get(key, 0.0)

    def set_qty(self, code, key, value):
        q = 0.0 if value in (None, "") else float(value)
        if q < 0:
            raise ValueError("qty")
        if q:
            self.clients[code]["qty"][key] = q
        else:
            self.clients[code]["qty"].pop(key, None)

    def amt(self, code, key):
        return self.clients[code]["amt"].get(key, 0.0)

    def set_amt(self, code, key, value):
        a = 0.0 if value in (None, "") else float(value)
        if a:
            self.clients[code]["amt"][key] = a
        else:
            self.clients[code]["amt"].pop(key, None)

    def line_total(self, code, key):
        """A priced line is quantity x rate. An unpriced line is whatever
        amount was typed for it. No usable level: nothing is priced."""
        if not self.priced(code):
            return 0.0
        r = self.rate(code, key)
        if r is None:
            return round(self.amt(code, key), 2)
        return round(self.qty(code, key) * float(r), 2)

    @staticmethod
    def piece_pick_key(weight):
        """Choose the outbound piece-pick rate bracket from weight per unit."""
        w = float(weight)
        if w < 1:
            return "o1"
        if w <= 5:
            return "o5"
        if w <= 10:
            return "o10"
        if w <= 20:
            return "o20"
        if w <= 40:
            return "o40"
        if w <= 50:
            return "o50"
        if w <= 70:
            return "o70"
        return "o99"

    @staticmethod
    def piece_receive_key(weight):
        """Choose the inbound receiving rate bracket from weight per unit."""
        w = float(weight)
        if w < 1: return "i1"
        if w <= 5: return "i5"
        if w <= 10: return "i10"
        if w <= 20: return "i20"
        if w <= 40: return "i40"
        if w <= 50: return "i50"
        if w <= 70: return "i70"
        return "i99"

    def add_pick(self, code, order_no, flow, mode, qty, weight, work_date, container_qty=0, sku_count=0):
        if not self.priced(code):
            raise ValueError("need_level")
        if flow == "inbound":
            key = self.piece_receive_key(weight)
            mode = "receive_piece"
        else:
            key = self.piece_pick_key(weight)
            mode = "piece"
        rate = self.rate(code, key)
        if rate is None:
            raise ValueError("need_pick_rate")
        record = {
            "order_no": order_no.strip(),
            "date": work_date,
            "flow": flow,
            "mode": mode,
            "qty": float(qty),
            "weight": float(weight),
            "container_qty": float(container_qty or 0),
            "sku_count": float(sku_count or 0),
            "key": key,
            "rate": float(rate),
            "amount": round(float(qty) * float(rate), 2),
        }
        self.clients[code].setdefault("picks", []).append(record)
        return record

    def remove_pick(self, code, index):
        picks = self.clients[code].setdefault("picks", [])
        if 0 <= index < len(picks):
            picks.pop(index)

    def update_pick(self, code, index, order_no, flow, mode, qty, weight, work_date, container_qty=0, sku_count=0):
        picks = self.clients[code].setdefault("picks", [])
        if not (0 <= index < len(picks)):
            raise ValueError("pick_item")
        old_len = len(picks)
        rec = self.add_pick(code, order_no, flow, mode, qty, weight, work_date, container_qty, sku_count)
        replacement = picks.pop()
        assert len(picks) == old_len
        picks[index] = replacement
        return replacement

    def pick_total(self, code):
        if not self.priced(code):
            return 0.0
        return round(sum(float(x.get("amount", 0))
                         for x in self.clients[code].get("picks", [])), 2)

    def group_total(self, code, group):
        if not self.priced(code):
            return 0.0
        base = sum(self.line_total(code, k)
                   for k in KEYS if CH[k]["grp"] == group)
        base += round(sum(float(x.get("amount", 0)) for x in self.clients[code].get("picks", []) if x.get("flow", "outbound") == group), 2)
        return round(base, 2)

    def total(self, code):
        return round(sum(self.line_total(code, k) for k in KEYS)
                     + self.pick_total(code), 2)

    def clear_period(self, code):
        self.clients[code]["qty"] = {}
        self.clients[code]["amt"] = {}
        self.clients[code]["picks"] = []

    def import_price_sheet(self, lid, path):
        """Read a price sheet into one level. Match each sheet row to a
        charge line. The weight labels repeat across Inbound / Outbound /
        Return, so the group heading in column A is the only thing that
        tells them apart."""
        import re
        from openpyxl import load_workbook
        if lid not in self.levels:
            raise ValueError("lv_none")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["Price"] if "Price" in wb.sheetnames else wb[wb.sheetnames[0]]

        by_group = {}
        for k in KEYS:
            if CH[k]["sheet"]:
                by_group.setdefault(CH[k]["grp"], {})[
                    CH[k]["sheet"].strip().lower()] = k
        racking = [k for k in KEYS if k.startswith("rk")]
        pallets = ["pltr", "pltl"]
        heads = {"inbound": "inbound", "outbound": "outbound",
                 "storage": "storage", "return": "return"}

        group, done, rack_i, plt_i = None, 0, 0, 0
        for r in ws.iter_rows(values_only=True):
            first = str(r[0]).strip().lower() if r and r[0] else ""
            if first in heads:
                group = heads[first]
            unit = str(r[1]).strip() if len(r) > 1 and r[1] else ""
            fee = r[2] if len(r) > 2 else None
            note = str(r[3]).strip().lower() if len(r) > 3 and r[3] else ""
            internal_key = str(r[4]).strip() if len(r) > 4 and r[4] else ""

            if not isinstance(fee, (int, float)):
                money = re.search(r"\$\s*(\d+(?:\.\d+)?)", str(fee))
                if not money:
                    continue
                fee = float(money.group(1))

            if internal_key in CH:
                self.set_level_rate(lid, internal_key, fee)
                done += 1
                continue

            if group == "storage" and not unit and "racking" in note:
                if rack_i < len(racking):
                    self.set_level_rate(lid, racking[rack_i], fee)
                    rack_i += 1
                    done += 1
                continue
            if unit.lower().startswith("pallet (charge used"):
                if plt_i < len(pallets):
                    self.set_level_rate(lid, pallets[plt_i], fee)
                    plt_i += 1
                    done += 1
                continue
            key = by_group.get(group, {}).get(unit.lower())
            if key:
                self.set_level_rate(lid, key, fee)
                done += 1
        if not done:
            raise ValueError("empty")
        return done

    # ---- file ------------------------------------------------------------

    def to_json(self):
        return json.dumps({"levels": self.levels, "clients": self.clients},
                          indent=2, ensure_ascii=False)

    @classmethod
    def from_json(cls, s, default_name="Default"):
        raw = json.loads(s)
        b = cls()
        for lid, lv in (raw.get("levels") or {}).items():
            if not isinstance(lv, dict):
                continue
            base = cls.fresh_rates()
            base.update(lv.get("rates") or {})
            b.levels[str(lid)] = {"name": str(lv.get("name") or lid),
                                  "rates": base}
        b.clients = raw.get("clients", {}) or {}
        fresh = cls.fresh_rates()
        for code, c in b.clients.items():
            c.setdefault("qty", {})
            c.setdefault("amt", {})
            c.setdefault("contact", "")
            c.setdefault("note", "")
            c["picks"] = [r for r in c.get("picks", [])
                          if r.get("mode") != "carton" and r.get("key") in CH]
            if "rates" in c:
                # 舊檔:單價掛在客戶上。搬到層級去 —— 單價完全一樣的客戶共用
                # 一層;跟報價表預設值一樣的那一層叫「預設」,其餘用客戶代碼
                # 命名,之後在 ① 改名即可。數量、單號都不動。
                base = cls.fresh_rates()
                base.update(c.pop("rates") or {})
                c["level"] = b._level_for(base, code, fresh, default_name)
            if not isinstance(c.get("level"), str):
                c["level"] = ""
        return b

    def _level_for(self, rates, code, fresh, default_name):
        for lid, lv in self.levels.items():
            if lv["rates"] == rates:
                return lid
        base_name = default_name if rates == fresh else code
        names = {lv["name"].lower() for lv in self.levels.values()}
        name, n = base_name, 2
        while name.lower() in names:
            name = f"{base_name} {n}"
            n += 1
        return self.add_level(name, rates)


class App:

    def __init__(self, root):
        self.root = root
        self.lang = "en"
        measure_ui_scale(root)
        resolve_fonts(root, chinese_ui=False)
        apply_ui_fonts()
        self.palette = self.load_theme()
        self.sketch = []
        self.b = self.load()
        self.cur = next(iter(sorted(self.b.clients)), None)
        # 正在編輯的層級:③ 費率頁與匯入/匯出報價表都作用在這一層。
        # 永遠至少有一層,新客戶才有東西可以指。
        self.edit_lv = self.b.ensure_level(self.tr("lv_default"))
        self._filling = False
        self._lost_lv = ""

        self.L, self.tkey = {}, {}
        self.themed, self.plain, self.cards = [], [], []
        self.rate_w, self.rate_rows, self.rate_grp = {}, [], {}
        self.trees = []
        self.pickers = []
        self._painting = False
        self.edit_pick_index = None

        root.geometry(f"{px(1150)}x{px(760)}")
        root.minsize(px(1120), px(720))
        self.s = ttk.Style()
        try:
            self.s.theme_use("clam")
        except tk.TclError:
            pass
        self.s.configure("TNotebook", borderwidth=0, padding=0,
                         tabmargins=(6, 6, 6, 0))
        self.s.configure("TNotebook.Tab", padding=(px(16), px(8)),
                         borderwidth=0, font=title_font(11, "bold"))
        row_font = tkfont.Font(font=data_font(10))
        self.s.configure("Treeview",
                         rowheight=max(px(28),
                                       row_font.metrics("linespace") + px(8)),
                         borderwidth=0, font=data_font(10))
        self.s.configure("Treeview.Heading", relief="flat",
                         font=title_font(10, "bold"), padding=(px(6), px(6)))

        self.head = tk.Frame(root)
        self.head.pack(fill="x", padx=px(18), pady=(px(12), 0))
        self.head_pet = SketchPet(self.head, kind="cat", size=px(42))
        self.head_pet.pack(side="left", padx=(0, px(8)))
        self.sketch.append(("pet", self.head_pet))
        self.t_app = tk.Label(self.head, font=title_font(20, "bold"))
        self.t_app.pack(side="left")
        self.t_sub = tk.Label(self.head, font=ui_font(10))
        self.t_sub.pack(side="left", padx=(px(10), 0), pady=(px(9), 0))
        self.lang_box = SketchCombo(self.head, values=["中文", "English"],
                                    width=8)
        self.lang_box.set("English")
        self.lang_box.pack(side="right")
        self.lang_box.bind("<<ComboboxSelected>>", self.on_lang)
        self.pal_box = SketchCombo(self.head, values=list(UI_THEMES))
        self.pal_box.pack(side="right", padx=px(6))
        self.pal_box.bind("<<ComboboxSelected>>", self.on_palette)
        self.sketch += [("combo", self.lang_box), ("combo", self.pal_box)]

        self.tools = tk.Frame(root)
        self.tools.pack(fill="x", padx=px(18), pady=(px(2), px(4)))
        self.plain.append(self.tools)
        self.rule = SketchRule(self.tools, width=px(240))
        self.rule.pack(side="left", fill="x", expand=True)
        self.sketch.append(("rule", self.rule))
        for key, cmd in (("save", self.save), ("exp", self.on_export_price_template), ("imp", self.on_import)):
            self.small(self.tools, key, cmd).pack(side="right", padx=(px(6), 0))

        self.body = tk.Frame(root)
        self.body.pack(fill="both", expand=True, padx=px(18))

        bar = tk.Frame(self.body)
        bar.pack(fill="x", pady=(0, px(6)))
        self.plain.append(bar)
        # 客戶不在這條列上:② ④ ⑤ 各自有搜尋框和清單,打字過濾、點一列
        # 就把客戶帶進那一頁。這裡只剩「正在編輯哪一層」—— 跟 UPS 對帳工具
        # 右上角那顆 Editing 一樣,每一頁都看得到,③ 費率頁改的就是它。
        self.lv_box = self.combo(bar, width=18)
        self.lv_box.pack(side="right")
        self.lv_box.bind("<<ComboboxSelected>>", self.on_pick_level)
        self.lab(bar, "l_editing", size=10, text_key="editing").pack(
            side="right", padx=(0, px(6)))

        self.nb = ttk.Notebook(self.body)
        self.nb.pack(fill="both", expand=True)
        self.pages = [tk.Frame(self.nb) for _ in range(5)]
        self.plain += self.pages
        for p in self.pages:
            self.nb.add(p, text="")
        self.build_levels(self.pages[0])
        self.build_clients(self.pages[1])
        self.build_rates(self.pages[2])
        self.build_calc(self.pages[3])
        self.build_bill(self.pages[4])

        self.foot = tk.Frame(root)
        self.foot.pack(fill="x", side="bottom", before=self.body)
        self.foot_in = tk.Frame(self.foot)
        self.foot_in.pack(fill="x", padx=px(18), pady=px(8))
        self.t_period = tk.Label(self.foot_in, font=ui_font(10))
        self.t_period.pack(side="left")
        self.total_val = tk.Label(self.foot_in, font=data_font(20, "bold"),
                                  text="$0.00")
        self.total_val.pack(side="left", padx=px(10))
        self.small(self.foot_in, "clear", self.on_clear, pop=POP["coral"]).pack(side="left")
        self.foot_pet = SketchPet(self.foot_in, kind="dog", size=px(42))
        self.foot_pet.pack(side="right", padx=(px(10), 0))
        self.sketch.append(("pet", self.foot_pet))
        self.status = tk.Label(self.foot_in, font=ui_font(10))
        self.status.pack(side="right")

        self.apply_palette()
        self.retext()
        self.select(self.cur)
        # 設定做過一次就不會再改;每個月打開來要做的事在 ④。
        if self.b.clients:
            self.nb.select(self.pages[3])
        self.maximize()

    def maximize(self):
        """開到最大。Windows 吃 zoomed,X11 吃 -zoomed;沒有視窗管理員
        (或兩個都不理)就直接填滿螢幕。minsize 留著,縮回去版面不會壞。"""
        root = self.root
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        for attempt in (lambda: root.state("zoomed"),
                        lambda: root.attributes("-zoomed", True)):
            try:
                attempt()
            except tk.TclError:
                continue
            root.update_idletasks()
            if root.winfo_width() >= sw - px(40):
                return
        root.geometry(f"{sw}x{sh}+0+0")

    def small(self, parent, key, cmd, primary=False, text_key=None,
              pop=None):
        """key identifies the widget; text_key names the string it shows.
        Two buttons may share a label -- registering both under one key
        silently dropped the first, which is how one ended up blank.

        主要動作是實心的蠟筆黃，其餘是線框；pop 給了就用那支蠟筆描邊，
        一排線框按鈕才分得出哪一顆會刪東西。"""
        if key in self.L:
            raise KeyError(f"duplicate widget key {key!r}")
        b = SketchButton(parent, command=cmd, font=title_font(10, "bold"),
                         padx=11, pady=5,
                         tone="solid" if primary else "soft", pop=pop)
        self.L[key] = b
        self.tkey[key] = text_key or key
        self.sketch.append(("button" if primary else "button2", b))
        return b

    def entry(self, parent, textvariable=None, width=10, **kw):
        e = SketchEntry(parent, textvariable=textvariable, width=width)
        for k in ("justify", "state"):
            if k in kw:
                e.configure(**{k: kw[k]})
        self.sketch.append(("entry", e))
        return e

    @staticmethod
    def span(text):
        return sum(2 if ord(ch) > 0x2E80 else 1 for ch in str(text))

    def fit(self, box, extra=3):
        values = box.cget("values") or ()
        if values:
            box.configure(width=max(self.span(v) for v in values) + extra)

    def combo(self, parent, values=(), width=None, textvariable=None):
        c = SketchCombo(parent, textvariable=textvariable, values=values,
                        width=width)
        self.sketch.append(("combo", c))
        return c

    def card(self, parent):
        """Audit Center 的 SketchPanel：2px 實線邊、奶油色面板。"""
        p = SketchPanel(parent, pad=2)
        self.sketch.append(("panel", p))
        inner = tk.Frame(p.body, padx=px(14), pady=px(12))
        inner.pack(fill="both", expand=True)
        self.cards.append(inner)
        return p, inner

    def lab(self, parent, key, kind="plain", size=9, text_key=None):
        w = tk.Label(parent, anchor="w", justify="left",
                     wraplength=px(940) if kind == "dim" else 0,
                     font=title_font(size + 1, "bold") if kind == "title"
                     else ui_font(size))
        if key in self.L:
            raise KeyError(f"duplicate widget key {key!r}")
        self.L[key] = w
        self.tkey[key] = text_key or key
        self.themed.append((kind, w))
        return w

    def page_head(self, page, n, scroll=False):
        """Every page says what it is for. Without it you are staring at a
        grid of boxes with no idea which one to fill."""
        host = page
        if scroll:
            outer, inner, cv, _bar = scrollable_page(page)
            outer.pack(fill="both", expand=True)
            self.plain += [outer, cv, inner]
            host = inner
        pad = tk.Frame(host)
        pad.pack(fill="both", expand=True, padx=px(10), pady=(px(12), px(10)))
        self.plain.append(pad)
        head = tk.Frame(pad)
        head.pack(fill="x", pady=(0, px(12)))
        self.plain.append(head)
        self.lab(head, f"p{n}_title", "title", 15).pack(anchor="w")
        self.lab(head, f"p{n}_help", "dim", 10).pack(anchor="w", pady=(px(2), 0))
        return pad

    # ---- client pickers ---------------------------------------------------

    def client_picker(self, parent, full=False, current=True, height=4):
        """A search box and a short list. Type to filter, click a row to
        bring that client into this page. One per page that works on a
        client -- there is no client box up top to go looking for."""
        n = len(self.pickers)
        box = tk.Frame(parent)
        box.pack(fill="both" if full else "x", expand=full, pady=(0, px(12)))
        self.plain.append(box)
        row = tk.Frame(box)
        row.pack(fill="x", pady=(0, px(6)))
        self.plain.append(row)
        self.lab(row, f"pk_l{n}", "plain", 10, text_key="find_client").pack(
            side="left", padx=(0, px(6)))
        var = tk.StringVar()
        self.entry(row, textvariable=var, width=26).pack(side="left")
        self.lab(row, f"pk_h{n}", "dim", 9, text_key="find_hint").pack(
            side="left", padx=(px(10), 0))
        cur = None
        if current:
            cur = tk.Label(row, font=title_font(12, "bold"), anchor="e")
            cur.pack(side="right")
            self.themed.append(("title", cur))
        if full:
            tv = self.table(box, (90, 200, 140, 130, 200), ("w",) * 5, height)
        else:
            tv = self.table(box, (90, 260, 160), ("w", "w", "w"), height,
                            expand=False)
        # 代碼欄固定寬,開到最大時多出來的寬度給名稱和備註,不是給代碼。
        tv.column("c0", stretch=False)
        pk = {"var": var, "tv": tv, "full": full, "cur": cur}
        self.pickers.append(pk)
        var.trace_add("write", lambda *a, pk=pk: self.fill_picker(pk))
        tv.bind("<<TreeviewSelect>>", lambda e, pk=pk: self.on_pick_row(pk))
        return pk

    def client_matches(self, code, q):
        d = self.b.clients[code]
        hay = " ".join((code, d["name"], d.get("contact", ""),
                        d.get("note", ""), self.level_label(code))).lower()
        return all(t in hay for t in q.lower().split())

    def current_label(self):
        if not self.cur:
            return self.tr("need_client")
        return f"{self.cur}  {self.b.clients[self.cur]['name']}"

    def fill_picker(self, pk):
        tv, q = pk["tv"], pk["var"].get().strip()
        codes = [c for c in sorted(self.b.clients)
                 if not q or self.client_matches(c, q)]
        self._filling = True
        try:
            tv.delete(*tv.get_children())
            for c in codes:
                d = self.b.clients[c]
                vals = (c, d["name"], self.level_label(c))
                if pk["full"]:
                    vals += (d.get("contact", ""), d.get("note", ""))
                tv.insert("", "end", iid=c, values=vals)
            self.stripe(tv)
            if self.cur in codes:
                tv.selection_set(self.cur)
                tv.see(self.cur)
        finally:
            self._filling = False
        if pk["cur"] is not None:
            pk["cur"].configure(text=self.current_label())

    def sync_pickers(self):
        """Every picker highlights the current client and says its name."""
        self._filling = True
        try:
            for pk in self.pickers:
                tv = pk["tv"]
                if self.cur and tv.exists(self.cur):
                    tv.selection_set(self.cur)
                    tv.see(self.cur)
                elif tv.selection():
                    tv.selection_remove(*tv.selection())
                if pk["cur"] is not None:
                    pk["cur"].configure(text=self.current_label())
        finally:
            self._filling = False

    @staticmethod
    def stripe(tv):
        """Alternate row shading. Identity tags stay first -- the calc page
        reads tags[0] to know which line a row is."""
        for i, iid in enumerate(tv.get_children()):
            tags = tv.item(iid, "tags")
            tags = [tags] if isinstance(tags, str) and tags else list(tags or ())
            tags = [t for t in tags if t not in ("odd", "even")]
            tv.item(iid, tags=tags + ["odd" if i % 2 else "even"])

    def duo_cards(self, host, left, right, wide_at=1500):
        """Two cards: side by side when the window is wide enough, stacked
        when it is not. Maximised they sit together; at the minimum size
        they stack, and neither layout squeezes a field."""
        host.columnconfigure(0, weight=1)
        host.columnconfigure(1, weight=1)
        state = {"wide": None}

        def relayout(_e=None):
            wide = host.winfo_width() >= px(wide_at)
            if wide == state["wide"]:
                return
            state["wide"] = wide
            for w in (left, right):
                w.grid_forget()
            if wide:
                left.grid(row=0, column=0, sticky="nsew", padx=(0, px(14)),
                          pady=(0, px(14)))
                right.grid(row=0, column=1, sticky="nsew", pady=(0, px(14)))
            else:
                left.grid(row=0, column=0, columnspan=2, sticky="ew",
                          pady=(0, px(14)))
                right.grid(row=1, column=0, columnspan=2, sticky="ew",
                           pady=(0, px(14)))
        host.bind("<Configure>", relayout)
        relayout()

    def money(self, v):
        return self.tr("money").format(v=f"{v:,.2f}")

    def tr(self, k):
        return T[self.lang][k]

    def name_of(self, k):
        return CH[k]["zh" if self.lang == "zh" else "en"]

    def unit_of(self, k):
        return CH[k]["u_zh" if self.lang == "zh" else "u_en"]

    def long_name(self, k):
        return f"{self.tr('grp')[CH[k]['grp']]} · {self.name_of(k)}"

    def say(self, m, flash=True):
        self.status.configure(text=m, fg=C["accent"] if flash else C["dim"])
        if flash:
            self.root.after(900, lambda: self.status.configure(fg=C["dim"]))

    def table(self, parent, widths, aligns, height, expand=True):
        panel = SketchPanel(parent, pad=2)
        panel.pack(fill="both" if expand else "x", expand=expand)
        self.sketch.append(("panel", panel))
        wrap = panel.body
        cols = tuple(f"c{i}" for i in range(len(widths)))
        tv = ttk.Treeview(wrap, columns=cols, show="headings", height=height)
        for c, w, a in zip(cols, widths, aligns):
            tv.column(c, width=px(w), anchor=a)
        sb = SketchScrollbar(wrap, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        tv.pack(side="left", fill="both", expand=True)
        self.trees.append(tv)
        return tv

    def build_levels(self, page):
        """One row per level. Pick a row and that is the level ③ edits;
        the top bar says so on every page."""
        pad = self.page_head(page, 1)
        form = tk.Frame(pad)
        form.pack(fill="x")
        self.plain.append(form)
        self.lv_name = tk.StringVar()
        cell = tk.Frame(form)
        cell.pack(side="left", padx=(0, 16))
        self.plain.append(cell)
        self.lab(cell, "f_lv_name", "dim", 9, text_key="lv_name").pack(anchor="w")
        self.entry(cell, textvariable=self.lv_name, width=24).pack()

        row = tk.Frame(pad)
        row.pack(fill="x", pady=(10, 12))
        self.plain.append(row)
        self.small(row, "lv_new", self.on_add_level, True).pack(side="left")
        self.small(row, "lv_copy", self.on_copy_level).pack(side="left", padx=px(6))
        self.small(row, "lv_rename", self.on_rename_level).pack(side="left")
        self.small(row, "lv_del", self.on_del_level, pop=POP["coral"]).pack(
            side="left", padx=px(6))

        row2 = tk.Frame(pad)
        row2.pack(fill="x", pady=(0, px(6)))
        self.plain.append(row2)
        self.lab(row2, "lv_list", "title", 10).pack(side="left")
        self.lv_q = tk.StringVar()
        self.entry(row2, textvariable=self.lv_q, width=22).pack(side="right")
        self.lab(row2, "f_lv_find", "plain", 10, text_key="find_level").pack(
            side="right", padx=(0, px(6)))
        self.lv_tv = self.table(pad, (260, 90, 110, 130),
                                ("w", "e", "e", "w"), 11)
        self.lv_tv.bind("<<TreeviewSelect>>", self.on_pick_level_row)
        self.lv_q.trace_add("write", lambda *a: self.fill_levels())

    def build_clients(self, page):
        pad = self.page_head(page, 2)
        form = tk.Frame(pad)
        form.pack(fill="x")
        self.plain.append(form)
        self.c_code, self.c_name = tk.StringVar(), tk.StringVar()
        self.c_contact, self.c_note = tk.StringVar(), tk.StringVar()
        for col, (key, var, w) in enumerate((("code", self.c_code, 10),
                                             ("name", self.c_name, 22),
                                             ("contact", self.c_contact, 14),
                                             ("note", self.c_note, 22))):
            cell = tk.Frame(form)
            cell.grid(row=0, column=col, sticky="w", padx=(0, 16))
            self.plain.append(cell)
            self.lab(cell, "f_" + key, "dim", 9, text_key=key).pack(anchor="w")
            self.entry(cell, textvariable=var, width=w).pack()
        # 層級跟代碼、名稱一起填,不是背後偷偷指定。新增時預設是正在編輯的
        # 那一層 —— 多半就是要指的那一層,而且它是看得見的。
        cell = tk.Frame(form)
        cell.grid(row=0, column=4, sticky="w")
        self.plain.append(cell)
        self.lab(cell, "f_level", "dim", 9, text_key="level").pack(anchor="w")
        self.c_level_box = self.combo(cell, width=14)
        self.c_level_box.pack()

        row = tk.Frame(pad)
        row.pack(fill="x", pady=(10, 12))
        self.plain.append(row)
        self.small(row, "cl_new", self.on_add, True).pack(side="left")
        self.small(row, "cl_edit", self.on_update).pack(side="left", padx=px(6))
        self.small(row, "del", self.on_del, pop=POP["coral"]).pack(side="left")

        self.lab(pad, "cl_list", "title", 10).pack(anchor="w", pady=(0, px(6)))
        self.pick_cus = self.client_picker(pad, full=True, current=False,
                                           height=11)
        self.cl_tv = self.pick_cus["tv"]

    def build_rates(self, page):
        pad = self.page_head(page, 3)
        # 這一頁改的是哪一層,以及有幾個客戶會跟著變。
        head = tk.Frame(pad)
        head.pack(fill="x", pady=(0, px(6)))
        self.plain.append(head)
        self.rate_who = tk.Label(head, font=title_font(12, "bold"), anchor="w")
        self.rate_who.pack(side="left")
        self.themed.append(("title", self.rate_who))
        self.rate_users = tk.Label(head, font=ui_font(9), anchor="w")
        self.rate_users.pack(side="left", padx=(px(10), 0), pady=(px(3), 0))
        self.themed.append(("dim", self.rate_users))
        row = tk.Frame(pad)
        row.pack(fill="x", pady=(0, 8))
        self.plain.append(row)
        self.lab(row, "l_group", "plain", 10, text_key="group").pack(side="left",
                                                                     padx=(0, px(6)))
        self.rt_grp = self.combo(row, width=12)
        self.rt_grp.pack(side="left")
        self.rt_grp.bind("<<ComboboxSelected>>", lambda e: self.show_rate_group())

        panel = SketchPanel(pad, pad=2)
        panel.pack(fill="both", expand=True)
        self.sketch.append(("panel", panel))
        outer, inner, cv, _bar = scrollable_page(panel.body)
        outer.pack(fill="both", expand=True)
        self.cards += [outer, cv, inner]
        self.rate_inner = inner

        for i, k in enumerate(KEYS):
            r = tk.Frame(inner)
            nm = tk.Label(r, font=ui_font(10), width=34, anchor="w")
            nm.pack(side="left", padx=(px(6), px(2)), pady=px(3))
            un = tk.Label(r, font=ui_font(9), width=9, anchor="w")
            un.pack(side="left")
            dollar = tk.Label(r, font=data_font(10))
            dollar.pack(side="left", padx=(0, px(2)))
            var = tk.StringVar()
            ent = self.entry(r, textvariable=var, width=10, justify="right")
            ent.pack(side="left")
            var.trace_add("write", lambda *a, kk=k: self.on_rate(kk))
            tag = tk.Label(r, font=ui_font(9), anchor="w")
            tag.pack(side="left", padx=px(8))
            self.rate_w[k] = (var, nm, un, ent, tag, dollar)
            self.rate_rows.append((i, k, r, nm, un, tag, dollar))

    def show_rate_group(self):
        """Only one group on screen at a time -- fifty-four rows down the page
        is a wall, not a form."""
        want = self.rt_grp.get()
        shown = 0
        for i, k, r, *_ in self.rate_rows:
            hit = want in ("", self.tr("all")) or \
                self.tr("grp")[CH[k]["grp"]] == want
            r.pack_forget()
            if hit:
                r.pack(fill="x")
                shown += 1
        self.zebra()
        return shown

    def zebra(self):
        n = 0
        for i, k, r, nm, un, tag, dollar in self.rate_rows:
            if not r.winfo_manager():
                continue
            bg = C["zebra"] if n % 2 else C["side"]
            for w in (r, nm, un, tag, dollar):
                w.configure(bg=bg)
            nm.configure(fg=C["ink"])
            un.configure(fg=C["dim"])
            tag.configure(fg=C["dim"])
            dollar.configure(fg=C["dim"])
            n += 1

    def build_calc(self, page):
        pad = self.page_head(page, 4, scroll=True)
        self.pick_calc = self.client_picker(pad)
        # 這個客戶用哪一層的單價 —— 或者為什麼算不出來。
        self.calc_lv = tk.Label(pad, font=ui_font(10, "bold"), anchor="w",
                                justify="left", wraplength=px(940))
        self.calc_lv.pack(anchor="w", pady=(0, px(8)))
        self.themed.append(("title", self.calc_lv))

        duo = tk.Frame(pad)
        duo.pack(fill="x")
        self.plain.append(duo)
        shell_pick, card = self.card(duo)
        self.pick_card = card

        self.lab(card, "pick_card", "title", 11).grid(
            row=0, column=0, columnspan=8, sticky="w", pady=(0, 3))
        self.lab(card, "pick_hint", "dim", 9).grid(
            row=1, column=0, columnspan=8, sticky="w", pady=(0, px(10)))

        from datetime import date
        today = date.today()
        self.order_var = tk.StringVar()
        self.flow_var = tk.StringVar()
        self.pick_mode_var = tk.StringVar()
        self.pick_qty_var = tk.StringVar()
        self.pick_weight_var = tk.StringVar()
        self.container_qty_var = tk.StringVar()
        self.sku_count_var = tk.StringVar()
        self.date_y = tk.StringVar(value=str(today.year))
        self.date_m = tk.StringVar(value=f"{today.month:02d}")
        self.date_d = tk.StringVar(value=f"{today.day:02d}")

        box = tk.Frame(card)
        box.grid(row=2, column=0, sticky="w", padx=(0, px(14)), pady=(0, px(6)))
        self.cards.append(box)
        self.lab(box, "pick_work_date", "dim", 9, text_key="work_date").pack(anchor="w")
        dr = tk.Frame(box); dr.pack(); self.cards.append(dr)
        self.date_y_box = self.combo(dr, textvariable=self.date_y, width=5, values=[str(y) for y in range(today.year-3, today.year+3)])
        self.date_m_box = self.combo(dr, textvariable=self.date_m, width=3, values=[f"{m:02d}" for m in range(1,13)])
        self.date_d_box = self.combo(dr, textvariable=self.date_d, width=3, values=[f"{d:02d}" for d in range(1,32)])
        for w in (self.date_y_box,self.date_m_box,self.date_d_box): w.pack(side="left", padx=(0,px(3)))

        fields = (("order_no", self.order_var, 12), ("flow_type", None, 8), ("pick_mode", None, 10), ("qty", self.pick_qty_var, 6), ("weight_each", self.pick_weight_var, 8), ("container_qty", self.container_qty_var, 6), ("sku_count", self.sku_count_var, 6))
        for i, (key, var, width) in enumerate(fields, start=1):
            box = tk.Frame(card)
            box.grid(row=2 + i // 4, column=i % 4, sticky="w",
                     padx=(0, px(14)), pady=(0, px(6)))
            self.cards.append(box)
            self.lab(box, "pick_" + key, "dim", 9, text_key=key).pack(anchor="w")
            if key == "flow_type":
                self.flow_box = self.combo(box, textvariable=self.flow_var, width=width)
                self.flow_box.pack(); self.flow_box.bind("<<ComboboxSelected>>", self.on_flow_change)
            elif key == "pick_mode":
                self.pick_mode_box = self.combo(box, textvariable=self.pick_mode_var, width=width)
                self.pick_mode_box.pack(); self.pick_mode_box.bind("<<ComboboxSelected>>", self.on_mode_change)
            else:
                ent = self.entry(box, textvariable=var, width=width, justify="right" if key in ("qty", "weight_each", "container_qty", "sku_count") else "left")
                ent.pack()
                if key == "weight_each": self.weight_entry = ent

        self.small(card, "add_pick", self.on_add_pick, True).grid(
            row=3, column=4, sticky="sw", padx=(px(4), 0), pady=(0, px(6)))

        shell_stor, sc = self.card(duo)
        self.storage_card = sc
        self.lab(sc, "storage_card", "title", 11).grid(row=0, column=0, columnspan=7, sticky="w", pady=(0, 3))
        self.lab(sc, "storage_hint", "dim", 9).grid(row=1, column=0, columnspan=7, sticky="w", pady=(0, px(10)))

        self.storage_start_var = tk.StringVar(value=today.strftime("%Y-%m-%d"))
        self.storage_end_var = tk.StringVar(value=today.strftime("%Y-%m-%d"))
        self.storage_cbm_var = tk.StringVar()
        self.storage_days_var = tk.StringVar(value="1")
        self.storage_rate_var = tk.StringVar(value="")
        self.storage_amount_var = tk.StringVar(value="$0.00")

        specs = (("storage_start", self.storage_start_var, 12),
                 ("storage_end", self.storage_end_var, 12),
                 ("storage_cbm", self.storage_cbm_var, 12))
        for col, (key, var, width) in enumerate(specs):
            bx = tk.Frame(sc); bx.grid(row=2, column=col, sticky="w", padx=(0, px(18))); self.cards.append(bx)
            self.lab(bx, "f_" + key, "dim", 9, text_key=key).pack(anchor="w")
            self.entry(bx, textvariable=var, width=width, justify="right" if key == "storage_cbm" else "left").pack()

        for col, (key, var, width) in enumerate((("storage_days", self.storage_days_var, 8),
                                                 ("storage_rate", self.storage_rate_var, 12),
                                                 ("storage_amount", self.storage_amount_var, 13))):
            bx = tk.Frame(sc)
            bx.grid(row=3, column=col, sticky="w", padx=(0, px(18)),
                    pady=(px(8), 0))
            self.cards.append(bx)
            self.lab(bx, "f_" + key, "dim", 9, text_key=key).pack(anchor="w")
            ent = self.entry(bx, textvariable=var, width=width, justify="right", state="readonly")
            ent.pack()

        self.small(sc, "add_storage", self.on_add_storage, True).grid(
            row=3, column=3, sticky="sw", padx=(px(4), 0), pady=(px(8), 0))
        for var in (self.storage_start_var, self.storage_end_var, self.storage_cbm_var):
            var.trace_add("write", lambda *a: self.refresh_storage_preview())
        self.duo_cards(duo, shell_pick, shell_stor)

        row = tk.Frame(pad)
        row.pack(fill="x", pady=(0, px(10)))
        self.plain.append(row)

        self.lab(row, "l_item", "plain", 10, text_key="item").pack(
            side="left", padx=(0, px(6)))
        self.it_box = self.combo(row, width=30)
        self.it_box.pack(side="left")
        self.it_box.bind("<<ComboboxSelected>>", lambda e: self.show_pick())

        self.lab(row, "l_qty", "plain", 10, text_key="qty").pack(
            side="left", padx=(px(14), px(6)))
        self.q_var = tk.StringVar()
        self.entry(row, textvariable=self.q_var, width=9,
                   justify="right").pack(side="left")
        self.q_unit = tk.Label(row, font=ui_font(9), width=6, anchor="w")
        self.q_unit.pack(side="left", padx=(px(4), 0))
        self.themed.append(("dim", self.q_unit))

        self.pick_rate = tk.Label(row, font=data_font(10), width=14,
                                  anchor="w")
        self.pick_rate.pack(side="left", padx=(px(10), 0))
        self.themed.append(("dim", self.pick_rate))
        self.small(row, "add", self.on_add_line, True).pack(
            side="left", padx=(px(10), 0))

        self.lab(pad, "added", "title", 10).pack(anchor="w", pady=(0, px(6)))
        self.calc_tv = self.table(
            pad, (125, 285, 75, 65, 100, 115),
            ("w", "w", "e", "w", "e", "e"), 11)
        row2 = tk.Frame(pad)
        row2.pack(fill="x", pady=(px(8), 0))
        self.plain.append(row2)
        self.small(row2, "rm", self.on_remove_line, pop=POP["coral"]).pack(side="left")
        self.small(row2, "edit_selected", self.on_edit_selected, True).pack(side="left", padx=px(6))
        self.small(row2, "cancel_edit", self.cancel_edit_mode).pack(side="left")
        self.L["cancel_edit"].configure(state="disabled")
        self.calc_tv.bind("<Double-1>", lambda e: self.on_edit_selected())

    def refresh_storage_preview(self):
        """Refresh days, rate and amount without interrupting partial typing."""
        try:
            from datetime import datetime
            start = datetime.strptime(self.storage_start_var.get().strip(), "%Y-%m-%d").date()
            end = datetime.strptime(self.storage_end_var.get().strip(), "%Y-%m-%d").date()
            days = (end - start).days + 1
            if days <= 0:
                raise ValueError
            cbm = float(self.storage_cbm_var.get().strip() or 0)
            rate = float(self.b.rate(self.cur, "stor") or 0) if self.cur else 0
            amount = max(cbm, 0) * days * rate
            self.storage_days_var.set(str(days))
            self.storage_rate_var.set(self.money(rate))
            self.storage_amount_var.set(self.money(amount))
        except Exception:
            self.storage_days_var.set("—")
            rate = float(self.b.rate(self.cur, "stor") or 0) if self.cur else 0
            self.storage_rate_var.set(self.money(rate))
            self.storage_amount_var.set(self.money(0))

    def on_add_storage(self):
        if not self.cur:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_client"))
        if not self.b.priced(self.cur):
            return messagebox.showwarning(self.tr("warn"), self.tr("need_level"))
        try:
            from datetime import datetime
            start = datetime.strptime(self.storage_start_var.get().strip(), "%Y-%m-%d").date()
            end = datetime.strptime(self.storage_end_var.get().strip(), "%Y-%m-%d").date()
            days = (end - start).days + 1
            if days <= 0:
                raise ValueError
        except Exception:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_storage_date"))
        try:
            cbm = float(self.storage_cbm_var.get().strip())
            if cbm <= 0:
                raise ValueError
        except ValueError:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_storage_cbm"))
        self.b.set_qty(self.cur, "stor", cbm * days)
        self.paint_calc(); self.paint_bill(); self.refresh_storage_preview()
        self.say(self.tr("ok"))

    def show_pick(self):
        """Show the chosen item's unit and rate right next to the box, so you
        can see what you are about to charge before adding it."""
        k = self.picked_key()
        if not k or not self.cur:
            self.q_unit.configure(text="")
            self.pick_rate.configure(text="")
            return
        self.q_unit.configure(text=self.unit_of(k))
        if not self.b.priced(self.cur):
            self.pick_rate.configure(text="—")
            return
        r = self.b.rate(self.cur, k)
        self.pick_rate.configure(
            text=self.tr("byquote") if r is None else "x " + self.money(r))

    def picked_key(self):
        i = self.it_box.current()
        return KEYS[i] if 0 <= i < len(KEYS) else None

    def build_bill(self, page):
        pad = self.page_head(page, 5)
        self.pick_bill = self.client_picker(pad, current=False)
        head = tk.Frame(pad)
        head.pack(fill="x")
        self.plain.append(head)
        self.bill_who = tk.Label(head, font=title_font(12, "bold"), anchor="w")
        self.bill_who.pack(side="left")
        self.themed.append(("title", self.bill_who))
        self.small(head, "exp_bill", self.on_export_bill, True).pack(side="right")
        self.small(head, "bill_tpl", self.on_export_template, pop=POP["mint"]).pack(side="right", padx=(0, px(6)))
        self.bill_lv = tk.Label(pad, font=ui_font(10, "bold"), anchor="w",
                                justify="left", wraplength=px(940))
        self.bill_lv.pack(anchor="w", pady=(px(2), px(6)))
        self.themed.append(("title", self.bill_lv))
        self.bill_tv = self.table(pad, (125, 285, 75, 65, 100, 115),
                                  ("w", "w", "e", "w", "e", "e"), 14)

    def load_theme(self):
        try:
            raw = THEME_FILE.read_text(encoding="utf-8").strip()
        except Exception:
            raw = ""
        parts = raw.split("|")
        if len(parts) < 2 or parts[0] != THEME_FILE_VERSION:
            return "Crayon"
        return parts[1] if parts[1] in UI_THEMES else "Crayon"

    def save_theme(self):
        try:
            THEME_FILE.write_text(f"{THEME_FILE_VERSION}|{self.palette}",
                                  encoding="utf-8")
        except Exception:
            pass

    def apply_palette(self):
        """跟 Audit Center 的 apply_theme 同一套角色：tk 元件逐顆補色，
        ttk 走 style，手繪殼走 retheme。"""
        global C
        C = palette(self.palette)
        t = UI_THEMES[self.palette]
        for w in (self.root, self.head, self.body, *self.plain):
            w.configure(bg=C["bg"])
        for w in (self.foot, self.foot_in, *self.cards):
            w.configure(bg=C["side"])
        self.t_app.configure(bg=C["bg"], fg=C["accent"])
        self.t_sub.configure(bg=C["bg"], fg=C["dim"])
        self.t_period.configure(bg=C["side"], fg=C["dim"])
        self.total_val.configure(bg=C["side"], fg=C["accent"])
        self.status.configure(bg=C["side"], fg=C["dim"])
        card_set = set(self.cards) | {self.foot_in}
        for kind, w in self.themed:
            try:
                parent = w.nametowidget(w.winfo_parent())
            except Exception:
                parent = None
            bg = C["side"] if parent in card_set else C["bg"]
            if kind == "title":
                w.configure(bg=bg, fg=C["accent"])
            elif kind == "dim":
                w.configure(bg=bg, fg=C["dim"])
            else:
                w.configure(bg=bg, fg=C["ink"])
        for kind, w in self.sketch:
            try:
                if kind == "button":
                    edge = t["LINE"] if _is_dark(t["RUN"]) else t["TITLE"]
                    w.retheme(page=t["BG"], fill=t["RUN"], outline=edge,
                              hover=t["RUN_HOVER"])
                elif kind == "button2":
                    w.retheme(page=t["BG"], fill=t["BG"], outline=t["LINE"],
                              hover=t["SELECT"], fg=t["TEXT"])
                elif kind == "entry":
                    w.retheme(page=t["BG"], panel=t["PANEL"], ink=t["TEXT"],
                              line=t["LINE"])
                elif kind == "combo":
                    w.retheme(page=t["BG"], panel=t["PANEL"], ink=t["TEXT"],
                              line=t["LINE"], hover=t["SELECT"])
                elif kind == "panel":
                    w.retheme(page=t["BG"], panel=t["PANEL"], line=t["LINE"])
                elif kind == "pet":
                    w.retheme(page=t["BG"], line=t["LINE"])
                elif kind == "rule":
                    w.configure(bg=t["BG"], fg=t["ACCENT"])
            except tk.TclError:
                pass
        try:
            self.foot_pet.retheme(page=t["PANEL"], line=t["LINE"])
        except tk.TclError:
            pass
        for bar in list(SketchScrollbar.instances):
            try:
                bar.retheme(page=t["BG"], trough=t["STRIPE"], line=t["LINE"],
                            thumb=t["TAB_BG"])
            except tk.TclError:
                SketchScrollbar.instances.remove(bar)
        self.zebra()
        self.s.configure("TNotebook", background=t["BG"])
        self.s.configure("TNotebook.Tab", background=t["TAB_BG"],
                         foreground=t["TEXT"])
        self.s.map("TNotebook.Tab",
                   background=[("selected", t["ACCENT"]),
                               ("active", t["SELECT"])],
                   foreground=[("selected", "white"), ("active", t["TEXT"])])
        self.s.configure("Treeview", background=t["PANEL"],
                         fieldbackground=t["PANEL"], foreground=t["TEXT"])
        self.s.configure("Treeview.Heading", background=t["ACCENT"],
                         foreground="white")
        self.s.map("Treeview.Heading",
                   background=[("active", t["RUN_HOVER"])])
        self.s.map("Treeview", background=[("selected", t["SELECT"])],
                   foreground=[("selected", t["TEXT"])])
        for tv in self.trees:
            tv.tag_configure("odd", background=t["STRIPE"])
            tv.tag_configure("even", background=t["PANEL"])

    def on_palette(self, _=None):
        name = self.pal_box.get()
        if name in UI_THEMES:
            self.palette = name
            self.save_theme()
        self.apply_palette()

    def on_lang(self, _=None):
        self.lang = "zh" if self.lang_box.get() == "中文" else "en"
        resolve_fonts(self.root, chinese_ui=self.lang == "zh")
        self.retext()
        self.select(self.cur)

    def on_add(self):
        level = self.level_from_box()
        try:
            code = self.b.add_client(self.c_code.get(), self.c_name.get(), level)
            self.b.update_client(code, self.c_name.get(), self.c_contact.get(),
                                 self.c_note.get())
        except ValueError as e:
            return messagebox.showwarning(self.tr("warn"), self.tr(str(e)))
        for v in (self.c_code, self.c_name, self.c_contact, self.c_note):
            v.set("")
        self.fill_levels()
        self.fill_clients()
        self.select(code)
        # 沒選層級不是錯,但它的作業算不出錢。現在就說,不要等到 ④。
        self.say(self.tr("ok") if self.b.priced(code) else self.tr("no_level"))

    def on_update(self):
        if not self.cur:
            return
        try:
            self.b.update_client(self.cur, self.c_name.get(),
                                 self.c_contact.get(), self.c_note.get(),
                                 self.level_from_box())
        except ValueError as e:
            return messagebox.showwarning(self.tr("warn"), self.tr(str(e)))
        self.fill_levels()
        self.fill_clients()
        self.select(self.cur)
        self.say(self.tr("ok") if self.b.priced(self.cur) else self.tr("no_level"))

    def on_del(self):
        if self.cur and messagebox.askyesno(self.tr("del"), self.cur):
            self.b.delete_client(self.cur)
            self.fill_levels()
            self.fill_clients()
            self.select(None)

    # ---- handlers: levels ------------------------------------------------

    def on_add_level(self):
        try:
            lid = self.b.add_level(self.lv_name.get())
        except ValueError as e:
            return messagebox.showwarning(self.tr("warn"), self.tr(str(e)))
        self.select_level(lid)
        self.say(self.tr("lv_added").format(name=self.b.level_name(lid)))

    def on_copy_level(self):
        if not self.edit_lv:
            return messagebox.showwarning(self.tr("warn"), self.tr("lv_none"))
        try:
            lid = self.b.copy_level(self.edit_lv, self.lv_name.get())
        except ValueError as e:
            return messagebox.showwarning(self.tr("warn"), self.tr(str(e)))
        self.select_level(lid)
        self.say(self.tr("lv_copied").format(name=self.b.level_name(lid)))

    def on_rename_level(self):
        if not self.edit_lv:
            return messagebox.showwarning(self.tr("warn"), self.tr("lv_none"))
        try:
            self.b.rename_level(self.edit_lv, self.lv_name.get())
        except ValueError as e:
            return messagebox.showwarning(self.tr("warn"), self.tr(str(e)))
        self.select_level(self.edit_lv)
        self.say(self.tr("ok"))

    def on_del_level(self):
        """Refuse before asking, and say who is on it -- so the person
        knows whom to move, instead of finding out at billing time."""
        lid = self.edit_lv
        if not lid:
            return
        name = self.b.level_name(lid)
        if len(self.b.levels) < 2:
            return messagebox.showwarning(self.tr("warn"), self.tr("lv_last"))
        users = self.b.level_users(lid)
        if users:
            return messagebox.showwarning(
                self.tr("warn"),
                self.tr("lv_inuse").format(name=name, n=len(users),
                                           who=self.tr("join").join(users)))
        if not messagebox.askyesno(self.tr("lv_del"),
                                   self.tr("lv_del_ask").format(name=name)):
            return
        try:
            self.b.delete_level(lid)
        except ValueError as e:
            return messagebox.showwarning(self.tr("warn"), self.tr(str(e)))
        self.select_level(None)
        self.say(self.tr("ok"))

    def on_pick_level(self, _=None):
        ids = self.b.level_ids()
        i = self.lv_box.current()
        if 0 <= i < len(ids):
            self.select_level(ids[i])

    def on_pick_level_row(self, _=None):
        """selection_set() in fill_levels() fires this too, asynchronously,
        so the row already being edited must be a no-op -- otherwise every
        repaint queues another select and the event loop never drains."""
        if self._filling:
            return
        sel = self.lv_tv.selection()
        if sel and sel[0] in self.b.levels and sel[0] != self.edit_lv:
            self.select_level(sel[0])

    def on_pick_row(self, pk, _=None):
        if self._filling:
            return
        sel = pk["tv"].selection()
        if sel and sel[0] in self.b.clients and sel[0] != self.cur:
            self.select(sel[0])

    def on_rate(self, key):
        """Every keystroke, on the level being edited. A half-typed number
        is ignored rather than nagging with a dialog; amounts update as soon
        as it parses -- for every client on this level."""
        lid = self.edit_lv
        if not lid or self._painting:
            return
        try:
            self.b.set_level_rate(lid, key,
                                  self.rate_w[key][0].get().strip() or None)
        except ValueError:
            return
        self.rate_w[key][4].configure(
            text=self.tr("byquote") if self.b.level_rate(lid, key) is None else "")
        if self.lv_tv.exists(lid):
            self.lv_tv.set(lid, "c2", f"{self.b.rates_set(lid)} / {len(KEYS)}")
        self.paint_calc()
        self.paint_bill()
        self.show_pick()
        if hasattr(self, "storage_start_var"):
            self.refresh_storage_preview()

    def on_flow_change(self, _=None):
        inbound = self.flow_box.current() == 0
        vals = [self.tr("receive_piece")] if inbound else [self.tr("piece_pick")]
        self.pick_mode_box.configure(values=vals)
        self.fit(self.pick_mode_box)
        self.pick_mode_box.current(0)
        self.on_mode_change()

    def on_mode_change(self, _=None):
        self.weight_entry.configure(state="normal")

    def on_add_pick(self):
        if not self.cur:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_client"))
        if not self.b.priced(self.cur):
            return messagebox.showwarning(self.tr("warn"), self.tr("need_level"))
        order_no = self.order_var.get().strip()
        if not order_no:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_order"))
        try:
            qty = float(self.pick_qty_var.get().strip())
            if qty <= 0:
                raise ValueError
        except ValueError:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_qty"))
        flow = "inbound" if self.flow_box.current() == 0 else "outbound"
        mode = "receive_piece" if flow == "inbound" else "piece"
        try:
            weight = float(self.pick_weight_var.get().strip())
            if weight <= 0: raise ValueError
        except ValueError:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_weight"))
        try:
            container_qty = float(self.container_qty_var.get().strip() or 0)
            sku_count = float(self.sku_count_var.get().strip() or 0)
            if container_qty < 0 or sku_count < 0:
                raise ValueError
        except ValueError:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_qty"))
        try:
            from datetime import date
            work_date = date(int(self.date_y.get()), int(self.date_m.get()), int(self.date_d.get())).isoformat()
        except Exception:
            return messagebox.showwarning(self.tr("warn"), self.tr("work_date"))
        try:
            if self.edit_pick_index is None:
                self.b.add_pick(self.cur, order_no, flow, mode, qty, weight, work_date, container_qty, sku_count)
            else:
                self.b.update_pick(self.cur, self.edit_pick_index, order_no, flow, mode, qty, weight, work_date, container_qty, sku_count)
        except ValueError as e:
            return messagebox.showwarning(self.tr("warn"), self.tr(str(e)))

        self.cancel_edit_mode(clear_fields=False)
        self.order_var.set("")
        self.pick_qty_var.set("")
        self.pick_weight_var.set("")
        self.container_qty_var.set("")
        self.sku_count_var.set("")
        self.paint_calc()
        self.paint_bill()
        self.say(self.tr("ok"))

    def on_edit_selected(self):
        sel = self.calc_tv.selection()
        if not sel or not self.cur:
            return
        tags = self.calc_tv.item(sel[0], "tags")
        if not tags:
            return
        tag = str(tags[0])
        if tag.startswith("pick:"):
            idx = int(tag.split(":", 1)[1])
            picks = self.b.clients[self.cur].get("picks", [])
            if not (0 <= idx < len(picks)):
                return
            rec = picks[idx]
            self.edit_pick_index = idx
            self.order_var.set(rec.get("order_no", ""))
            self.pick_qty_var.set(f"{float(rec.get('qty', 0)):g}")
            self.pick_weight_var.set(f"{float(rec.get('weight', 0)):g}")
            self.container_qty_var.set(f"{float(rec.get('container_qty', 0)):g}" if rec.get("container_qty") else "")
            self.sku_count_var.set(f"{float(rec.get('sku_count', 0)):g}" if rec.get("sku_count") else "")
            try:
                y, m, d = str(rec.get("date", "")).split("-")
                self.date_y.set(y); self.date_m.set(m); self.date_d.set(d)
            except ValueError:
                pass
            self.flow_box.current(0 if rec.get("flow") == "inbound" else 1)
            self.on_flow_change()
            self.buttons_edit_state(True)
            return

        key = tag
        try:
            idx = KEYS.index(key)
        except ValueError:
            return
        self.it_box.current(idx)
        value = self.b.amt(self.cur, key) if self.b.rate(self.cur, key) is None else self.b.qty(self.cur, key)
        self.q_var.set(f"{float(value):g}")
        self.show_pick()

    def buttons_edit_state(self, editing):
        self.L["add_pick"].configure(text=self.tr("save_edit") if editing else self.tr("add_pick"))
        self.L["cancel_edit"].configure(state="normal" if editing else "disabled")

    def cancel_edit_mode(self, clear_fields=True):
        self.edit_pick_index = None
        if clear_fields:
            self.order_var.set("")
            self.pick_qty_var.set("")
            self.pick_weight_var.set("")
        self.buttons_edit_state(False)

    def on_add_line(self):
        if not self.cur:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_client"))
        if not self.b.priced(self.cur):
            return messagebox.showwarning(self.tr("warn"), self.tr("need_level"))
        k = self.picked_key()
        if not k:
            return messagebox.showwarning(self.tr("warn"), self.tr("pick_item"))
        txt = self.q_var.get().strip()
        try:
            v = float(txt)
            if v <= 0:
                raise ValueError
        except ValueError:
            unpriced = self.b.rate(self.cur, k) is None
            return messagebox.showwarning(
                self.tr("warn"),
                self.tr("need_amt") if unpriced else self.tr("need_qty"))
        if self.b.rate(self.cur, k) is None:
            self.b.set_amt(self.cur, k, v)
        else:
            self.b.set_qty(self.cur, k, v)
        self.q_var.set("")
        self.paint_calc()
        self.paint_bill()
        self.say(self.tr("ok"))

    def on_remove_line(self):
        sel = self.calc_tv.selection()
        if not sel or not self.cur:
            return
        tags = self.calc_tv.item(sel[0], "tags")
        if not tags:
            return
        tag = tags[0]
        if str(tag).startswith("pick:"):
            idx = int(str(tag).split(":", 1)[1])
            self.b.remove_pick(self.cur, idx)
            self.cancel_edit_mode()
        else:
            self.b.set_qty(self.cur, tag, 0)
            self.b.set_amt(self.cur, tag, 0)
        self.paint_calc()
        self.paint_bill()
        self.say(self.tr("ok"))

    def on_clear(self):
        if self.cur and messagebox.askyesno(self.tr("clear"),
                                            self.tr("clear_ask")):
            self.b.clear_period(self.cur)
            self.paint_calc()
            self.paint_bill()
            self.say(self.tr("ok"))

    def on_import(self):
        """Read a price sheet into the level being edited -- the one named
        in the top bar, so the file lands where you can see it will."""
        lid = self.edit_lv
        if not lid:
            return messagebox.showwarning(self.tr("warn"), self.tr("lv_none"))
        try:
            import openpyxl
        except ImportError:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_xl"))
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm"),
                                                  ("All", "*.*")])
        if not p:
            return
        try:
            n = self.b.import_price_sheet(lid, p)
        except ValueError:
            return messagebox.showwarning(self.tr("warn"), self.tr("imp_fail"))
        except Exception as e:
            return messagebox.showerror(self.tr("warn"), str(e))
        self.select_level(lid)
        self.say(self.tr("imp_ok").format(n=n, lv=self.b.level_name(lid)))

    @staticmethod
    def safe_name(text):
        return "".join(ch if ch.isalnum() or ch in "-_" else "_"
                       for ch in str(text)).strip("_") or "level"

    def on_export_price_template(self):
        """Export the level being edited as a price sheet that can be
        imported again -- into this level or another one."""
        lid = self.edit_lv
        if not lid:
            return messagebox.showwarning(self.tr("warn"), self.tr("lv_none"))
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_xl"))
        lv_name = self.b.level_name(lid)
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=f"Geniqua_Price_Sheet_{self.safe_name(lv_name)}.xlsx", filetypes=[("Excel", "*.xlsx")])
        if not p:
            return
        wb = Workbook(); ws = wb.active; ws.title = "Price"
        ws.append([f"GENIQUA LOGISTICS - PRICE SHEET - {lv_name}", "", "", ""])
        ws.append(["Category", "Price Sheet Item / Unit", "Fee", "Notes"])
        labels={"inbound":"INBOUND","outbound":"OUTBOUND","storage":"STORAGE","return":"RETURN"}
        for g in GROUPS:
            ws.append([labels[g], "", "", ""])
            for k in KEYS:
                if CH[k]["grp"] != g:
                    continue
                rate=self.b.level_rate(lid,k)
                ws.append(["", CH[k].get("sheet") or self.name_of(k), "" if rate is None else rate, f"{CH[k]['zh']} / {CH[k]['en']} | Unit: {CH[k]['u_en']}"])
        ws.merge_cells("A1:D1")
        ws["A1"].font=Font(bold=True,size=15,color="FFFFFF"); ws["A1"].fill=PatternFill("solid",fgColor="0E7C86"); ws["A1"].alignment=Alignment(horizontal="center")
        soft=PatternFill("solid",fgColor="EAF1F4"); line=Side(style="thin",color="D9E3E7")
        for c in ws[2]: c.fill=soft; c.font=Font(bold=True); c.alignment=Alignment(horizontal="center")
        for r in range(3,ws.max_row+1):
            if ws.cell(r,1).value in labels.values():
                for c in ws[r]: c.fill=PatternFill("solid",fgColor="D2E7EA"); c.font=Font(bold=True,color="0E7C86")
            else: ws.cell(r,3).number_format='$#,##0.00'
            for c in ws[r]: c.border=Border(bottom=line); c.alignment=Alignment(vertical="center",wrap_text=True)
        ws.column_dimensions["A"].width=16; ws.column_dimensions["B"].width=42; ws.column_dimensions["C"].width=14; ws.column_dimensions["D"].width=58
        ws.auto_filter.ref=f"A2:D{ws.max_row}"; ws.sheet_view.showGridLines=False
        wb.save(p); messagebox.showinfo(self.tr("ok"), self.tr("export_ok").format(path=p))

    def _style_sheet(self, ws, cols):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        accent="0E7C86"; soft="EAF1F4"; line=Side(style="thin", color="D9E3E7")
        for c in ws[1]: c.font=Font(bold=True,size=14,color=accent)
        for c in ws[2]: c.fill=PatternFill("solid",fgColor=soft); c.font=Font(bold=True); c.alignment=Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            for c in row: c.border=Border(bottom=line); c.alignment=Alignment(vertical="center")
        widths=[16,38,14,14,14,16,16,16]
        for i in range(1, cols+1): ws.column_dimensions[chr(64+i)].width=widths[i-1]
        ws.auto_filter.ref=ws.dimensions

    def on_export_template(self):
        try:
            from openpyxl import Workbook
        except ImportError:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_xl"))
        p=filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="warehouse_invoice_template.xlsx", filetypes=[("Excel","*.xlsx")])
        if not p:return
        wb=Workbook(); ws=wb.active; ws.title="Invoice"
        ws.append(["GENIQUA LOGISTICS - WAREHOUSE BILLING"]); ws.append(["Date","Order No.","Operation","Billing Method","Item","Qty","Weight (lb)","Unit","Rate","Amount"])
        for _ in range(8): ws.append([""]*10)
        ws.append(["","","","","","","","","TOTAL",""])
        self._style_sheet(ws,10); ws.merge_cells("A1:J1"); wb.save(p); self.say(self.tr("ok"))

    def on_export_bill(self):
        """Export a compact Geniqua-style warehouse invoice."""
        if not self.cur:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_client"))
        if not self.b.priced(self.cur):
            return messagebox.showwarning(self.tr("warn"), self.tr("need_level"))
        c=self.b.clients[self.cur]
        picks=[r for r in c.get("picks",[]) if r.get("key") in CH]
        regular=[]
        for g in GROUPS:
            for k in KEYS:
                amount=self.b.line_total(self.cur,k)
                if CH[k]["grp"]==g and amount: regular.append((g,k,amount))
        if not picks and not regular:
            return messagebox.showwarning(self.tr("warn"), self.tr("no_bill_data"))
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
            from datetime import date,datetime
            from openpyxl.utils import get_column_letter
        except ImportError:
            return messagebox.showwarning(self.tr("warn"), self.tr("need_xl"))
        today=date.today(); invoice_no=f"GQL-{today:%m%d%y}-{self.cur}"
        p=filedialog.asksaveasfilename(defaultextension=".xlsx",initialfile=f"{invoice_no}.xlsx",filetypes=[("Excel","*.xlsx")])
        if not p:return
        wb=Workbook(); ws=wb.active; ws.title="INVOICE"
        ws.append(["GENIQUA LOGISTICS - WAREHOUSE INVOICE"]+[""]*8)
        ws.append(["Client",c["name"],"Invoice No.",invoice_no,"Invoice Date",today,"Rate Level",self.level_label(self.cur),""])
        headers=["Date","Order No.","Operation","Billing Item","Containers","SKU Count","Qty","Unit Price","Amount"]
        ws.append(headers)
        for rec in picks:
            raw=rec.get("date","")
            try: d=datetime.strptime(raw,"%Y-%m-%d").date()
            except Exception: d=raw
            flow="Inbound" if rec.get("flow")=="inbound" else "Outbound"
            ws.append([d,rec.get("order_no",""),flow,self.name_of(rec["key"]),rec.get("container_qty",0) or "",rec.get("sku_count",0) or "",rec.get("qty",0),rec.get("rate",0),rec.get("amount",0)])
        for g,k,amount in regular:
            r=self.b.rate(self.cur,k)
            ws.append([today,"",self.tr("grp")[g],self.name_of(k),"","",self.b.qty(self.cur,k),"" if r is None else r,amount])
        data_end=ws.max_row
        ws.append(["","","","TOTAL","","","","",f"=SUM(I4:I{data_end})"])
        ws.merge_cells("A1:I1"); ws["A1"].font=Font(bold=True,size=17,color="FFFFFF"); ws["A1"].fill=PatternFill("solid",fgColor="0E7C86"); ws["A1"].alignment=Alignment(horizontal="center")
        ws["F2"].number_format="mm/dd/yyyy"
        fill=PatternFill("solid",fgColor="D9E3E7"); line=Side(style="thin",color="C7D3D9")
        for c0 in ws[3]: c0.fill=fill; c0.font=Font(bold=True); c0.alignment=Alignment(horizontal="center"); c0.border=Border(top=line,bottom=line)
        for row in ws.iter_rows(min_row=4,max_row=ws.max_row):
            for c0 in row: c0.border=Border(bottom=Side(style="hair",color="D9E3E7")); c0.alignment=Alignment(vertical="center",wrap_text=True)
            row[0].number_format="mm/dd/yyyy"; row[7].number_format='$#,##0.00'; row[8].number_format='$#,##0.00'
        for c0 in ws[ws.max_row]: c0.font=Font(bold=True); c0.fill=PatternFill("solid",fgColor="EAF1F4")
        widths=[12,20,13,38,12,12,10,13,14]
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
        ws.auto_filter.ref=f"A3:I{data_end}"; ws.sheet_view.showGridLines=False
        ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1
        self.save(); wb.save(p); messagebox.showinfo(self.tr("ok"),self.tr("export_ok").format(path=p)); self.say(self.tr("ok"))

    def retext(self):
        for k, w in self.L.items():
            w.configure(text=self.tr(self.tkey.get(k, k)))
        self.t_app.configure(text=self.tr("app"))
        self.t_sub.configure(text=self.tr("sub"))
        self.root.title(self.tr("app"))
        self.t_period.configure(text=self.tr("period"))
        for i, name in enumerate(self.tr("pages")):
            self.nb.tab(self.pages[i], text=f"  {name}  ")
        for pk in self.pickers:
            cols = self.tr("cl_cols") if pk["full"] else self.tr("pick_cols")
            for i, h in enumerate(cols):
                pk["tv"].heading(f"c{i}", text=h)
        for cid, h in zip(("c0", "c1", "c2", "c3"), self.tr("lv_cols")):
            self.lv_tv.heading(cid, text=h)
        for tv in (self.calc_tv, self.bill_tv):
            for cid, h in zip(("c0", "c1", "c2", "c3", "c4", "c5"),
                              self.tr("calc_cols")):
                tv.heading(cid, text=h)
        for i, k, r, nm, un, tag, dollar in self.rate_rows:
            nm.configure(text=self.name_of(k))
            un.configure(text=self.unit_of(k))
            dollar.configure(text="$")
        groups = [self.tr("all")] + [self.tr("grp")[g] for g in GROUPS]
        self.rt_grp.configure(values=groups)
        if self.rt_grp.get() not in groups:
            self.rt_grp.set(groups[1])
        self.it_box.configure(values=[self.long_name(k) for k in KEYS])
        self.flow_box.configure(values=[self.tr("inbound_flow"), self.tr("outbound_flow")])
        self.fit(self.flow_box)
        if self.flow_box.current() < 0: self.flow_box.current(0)
        self.on_flow_change()
        if self.it_box.current() < 0:
            self.it_box.current(0)
        self.pal_box.set(self.palette)
        for box in (self.rt_grp, self.pick_mode_box):
            self.fit(box)
        self.show_rate_group()
        self.fill_levels()
        self.fill_clients()
        self.paint()

    # ---- levels on screen -------------------------------------------------

    def level_label(self, code):
        """The level a client is on, by name -- or why it has none."""
        lid, status = self.b.resolve(code)
        if status == "ok":
            return self.b.level_name(lid)
        if status == "lost":
            return self.tr("gone").format(id=lid)
        return self.tr("unset")

    def level_note(self, code):
        lid, status = self.b.resolve(code)
        if status == "ok":
            return self.tr("lv_of").format(name=self.b.level_name(lid))
        if status == "lost":
            return self.tr("lost_level").format(id=lid)
        return self.tr("no_level")

    def fill_levels(self):
        """Top-bar box, the ① table and the client form's level box all list
        the same levels in the same order; this is the one place they are
        filled, so they cannot drift apart."""
        ids = self.b.level_ids()
        names = [self.b.level_name(l) for l in ids]
        self._filling = True
        try:
            self.lv_box.configure(values=names)
            self.fit(self.lv_box)
            if self.edit_lv in ids:
                self.lv_box.current(ids.index(self.edit_lv))
            else:
                self.lv_box.set("")
            q = self.lv_q.get().strip().lower()
            shown = [l for l in ids
                     if not q or q in self.b.level_name(l).lower()]
            self.lv_tv.delete(*self.lv_tv.get_children())
            for lid in shown:
                self.lv_tv.insert("", "end", iid=lid, values=(
                    self.b.level_name(lid), len(self.b.level_users(lid)),
                    f"{self.b.rates_set(lid)} / {len(KEYS)}",
                    self.tr("editing") if lid == self.edit_lv else ""))
            self.stripe(self.lv_tv)
            if self.edit_lv in shown:
                self.lv_tv.selection_set(self.edit_lv)
                self.lv_tv.see(self.edit_lv)
            self.set_level_box(self.cur)
        finally:
            self._filling = False

    def set_level_box(self, code):
        """The client form's level box. A client pointing at a level that no
        longer exists keeps that entry, spelled out as deleted -- otherwise
        opening the form once would silently clear it."""
        ids = self.b.level_ids()
        vals = [self.tr("unset")] + [self.b.level_name(l) for l in ids]
        lid, status = self.b.resolve(code) if code else ("", "none")
        self._lost_lv = lid if status == "lost" else ""
        if status == "lost":
            vals.append(self.tr("gone").format(id=lid))
        self.c_level_box.configure(values=vals)
        self.fit(self.c_level_box)
        if status == "ok":
            self.c_level_box.current(ids.index(lid) + 1)
        elif status == "lost":
            self.c_level_box.current(len(vals) - 1)
        elif code:
            self.c_level_box.current(0)
        elif self.edit_lv in ids:
            # 空白表單(要新增):預設落在正在編輯的那一層。
            self.c_level_box.current(ids.index(self.edit_lv) + 1)
        else:
            self.c_level_box.current(0)

    def level_from_box(self):
        ids = self.b.level_ids()
        i = self.c_level_box.current()
        if 1 <= i <= len(ids):
            return ids[i - 1]
        if i == len(ids) + 1:
            return self._lost_lv
        return ""

    def select_level(self, lid):
        """Make lid the level being edited and repaint everything that
        shows a level name or a price."""
        ids = self.b.level_ids()
        self.edit_lv = lid if lid in self.b.levels else (ids[0] if ids else None)
        if self.edit_lv:
            self.lv_name.set(self.b.level_name(self.edit_lv))
        self.fill_levels()
        self.fill_clients()
        self.paint()

    def fill_clients(self):
        for pk in self.pickers:
            self.fill_picker(pk)

    def select(self, code):
        codes = sorted(self.b.clients)
        self.cur = code if code in self.b.clients else (codes[0] if codes else None)
        if self.cur:
            d = self.b.clients[self.cur]
            self.c_code.set(self.cur)
            self.c_name.set(d["name"])
            self.c_contact.set(d.get("contact", ""))
            self.c_note.set(d.get("note", ""))
            self.set_level_box(self.cur)
            self.say(f"{self.cur}  {d['name']} · {self.level_label(self.cur)}",
                     False)
        else:
            self.set_level_box(None)
            self.say(self.tr("need_client"), False)
        # 每一頁的清單都要跟著 cur 走,不然剛填表時選到的舊列會在下一輪事件
        # 裡把 cur 再切回去。
        self.sync_pickers()
        self.paint()

    def paint(self):
        self.paint_rates()
        self.paint_calc()
        self.paint_bill()
        self.show_pick()

    def paint_rates(self):
        lid = self.edit_lv if self.edit_lv in self.b.levels else None
        if lid:
            n = len(self.b.level_users(lid))
            self.rate_who.configure(
                text=self.tr("rate_for").format(name=self.b.level_name(lid)))
            self.rate_users.configure(
                text=self.tr("lv_users").format(n=n) if n else self.tr("lv_nobody"))
        else:
            self.rate_who.configure(text=self.tr("lv_none"))
            self.rate_users.configure(text="")
        self._painting = True
        for k, (var, _nm, _un, ent, tag, _d) in self.rate_w.items():
            if not lid:
                var.set("")
                ent.configure(state="disabled")
                tag.configure(text="")
                continue
            ent.configure(state="normal")
            r = self.b.level_rate(lid, k)
            var.set("" if r is None else f"{r:g}")
            tag.configure(text=self.tr("byquote") if r is None else "")
        self._painting = False

    def paint_calc(self):
        self.calc_tv.delete(*self.calc_tv.get_children())
        if not self.cur:
            self.calc_lv.configure(text="")
            self.total_val.configure(text=self.money(0))
            return
        self.calc_lv.configure(text=self.level_note(self.cur))
        if not self.b.priced(self.cur):
            # 沒有可用的層級就沒有單價。已填的量照列,單價與金額留白,
            # 合計是「—」—— 不要顯示一個少算了的數字。
            c = self.b.clients[self.cur]
            for i, rec in enumerate(c.get("picks", [])):
                k = rec["key"]
                self.calc_tv.insert("", "end", tags=(f"pick:{i}",), values=(
                    f"{rec.get('date','')}  {rec.get('order_no','')}",
                    self.name_of(k), f"{rec.get('qty', 0):g}",
                    self.unit_of(k), "—", "—"))
            for k in KEYS:
                q, a = c["qty"].get(k, 0), c["amt"].get(k, 0)
                if not q and not a:
                    continue
                self.calc_tv.insert("", "end", tags=(k,), values=(
                    "", self.long_name(k), f"{q:g}" if q else "",
                    self.unit_of(k), "—", self.money(a) if a else "—"))
            self.stripe(self.calc_tv)
            self.total_val.configure(text="—")
            return

        for i, rec in enumerate(self.b.clients[self.cur].get("picks", [])):
            k = rec["key"]
            self.calc_tv.insert("", "end", tags=(f"pick:{i}",), values=(
                f"{rec.get('date','')}  {rec.get('order_no','')}", self.name_of(k) + ((f" | 櫃 {rec.get('container_qty'):g} | SKU {rec.get('sku_count'):g}") if rec.get("flow")=="inbound" and (rec.get("container_qty") or rec.get("sku_count")) else ""),
                f"{rec.get('qty', 0):g}", self.unit_of(k),
                self.money(rec.get("rate", 0)), self.money(rec.get("amount", 0))))

        for k in KEYS:
            a = self.b.line_total(self.cur, k)
            if not a:
                continue
            r = self.b.rate(self.cur, k)
            self.calc_tv.insert("", "end", tags=(k,), values=(
                "", self.long_name(k), f"{self.b.qty(self.cur, k):g}",
                self.unit_of(k),
                self.tr("byquote") if r is None else self.money(r),
                self.money(a)))
        self.stripe(self.calc_tv)
        self.total_val.configure(text=self.money(self.b.total(self.cur)))

    def paint_bill(self):
        self.bill_tv.delete(*self.bill_tv.get_children())
        if not self.cur:
            self.bill_who.configure(text="")
            self.bill_lv.configure(text="")
            return
        c = self.b.clients[self.cur]
        self.bill_who.configure(text=f"{self.cur}   {c['name']}")
        self.bill_lv.configure(text=self.level_note(self.cur))
        if not self.b.priced(self.cur):
            return
        for g in GROUPS:
            regular = [k for k in KEYS if CH[k]["grp"] == g
                       and self.b.line_total(self.cur, k)]
            picks = c.get("picks", []) if g == "outbound" else []
            if not regular and not picks:
                continue
            self.bill_tv.insert("", "end", values=(
                "", f"— {self.tr('grp')[g]} —", "", "", "", ""))
            for rec in picks:
                k = rec["key"]
                self.bill_tv.insert("", "end", values=(
                    rec.get("order_no", ""), "   " + self.name_of(k),
                    f"{rec.get('qty', 0):g}", self.unit_of(k),
                    self.money(rec.get("rate", 0)),
                    self.money(rec.get("amount", 0))))
            for k in regular:
                r = self.b.rate(self.cur, k)
                self.bill_tv.insert("", "end", values=(
                    "", "   " + self.name_of(k),
                    f"{self.b.qty(self.cur, k):g}", self.unit_of(k),
                    self.tr("byquote") if r is None else self.money(r),
                    self.money(self.b.line_total(self.cur, k))))
            self.bill_tv.insert("", "end", values=(
                "", self.tr("subtotal").format(g=self.tr("grp")[g]),
                "", "", "", self.money(self.b.group_total(self.cur, g))))
        self.bill_tv.insert("", "end", values=(
            "", self.tr("period"), "", "", "", self.money(self.b.total(self.cur))))
        self.stripe(self.bill_tv)

    def load(self):
        if DATA.exists():
            try:
                return Book.from_json(DATA.read_text(encoding="utf-8"),
                                      default_name=self.tr("lv_default"))
            except Exception:
                pass
        return Book()

    def save(self):
        try:
            DATA.write_text(self.b.to_json(), encoding="utf-8")
            self.say(self.tr("saved"))
        except Exception as e:
            messagebox.showerror(self.tr("warn"), str(e))


def main():
    enable_dpi_awareness()
    root = tk.Tk()
    app = App(root)
    root.protocol("WM_DELETE_WINDOW", lambda: (app.save(), root.destroy()))
    root.mainloop()


if __name__ == "__main__":
    main()
